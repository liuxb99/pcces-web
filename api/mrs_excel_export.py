"""Phase 3 Legacy-compatible MRS Excel export.

The workbook keeps the two-grid semantics used by FormBudgetRes: a resource
summary sheet and a resource-to-budget-item reference sheet. Decimal display
formats follow the project precision policy instead of a global scale.
"""
from __future__ import annotations

from io import BytesIO

from flask import Blueprint, Response, jsonify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import and_, select

from api.budget_decimal import budget_items_decimal
from api.mrs_precision_policy import MRSPrecisionPolicyService
from api.resource_budget_lineage import resource_budget_links
from api.resource_decimal import resources_decimal

RESOURCE_HEADERS = ["資源編碼", "資源名稱", "單位", "單價", "引用工項數"]
REFERENCE_HEADERS = ["資源編碼", "工項編號", "工項名稱", "數量", "單價", "金額"]


class MRSExcelExportService:
    def __init__(self, engine):
        self.engine = engine
        self.precision = MRSPrecisionPolicyService(engine)

    @staticmethod
    def _fmt(scale: int) -> str:
        return "0" if scale == 0 else "0." + ("0" * scale)

    def export_project(self, project_code: str) -> bytes:
        policy = self.precision.get(project_code)
        with self.engine.connect() as conn:
            resource_rows = conn.execute(
                select(resources_decimal.c.id, resources_decimal.c.code, resources_decimal.c.name,
                       resources_decimal.c.unit, resources_decimal.c.unit_price)
                .join(resource_budget_links, resource_budget_links.c.resource_id == resources_decimal.c.id)
                .where(resource_budget_links.c.project_code == project_code)
                .distinct().order_by(resources_decimal.c.code)
            ).mappings().all()
            refs = conn.execute(
                select(resources_decimal.c.code.label("resource_code"), budget_items_decimal)
                .join(resource_budget_links, resource_budget_links.c.resource_id == resources_decimal.c.id)
                .join(budget_items_decimal, budget_items_decimal.c.id == resource_budget_links.c.budget_item_id)
                .where(and_(resource_budget_links.c.project_code == project_code,
                            budget_items_decimal.c.project_code == project_code))
                .order_by(resources_decimal.c.code, budget_items_decimal.c.item_no, budget_items_decimal.c.id)
            ).mappings().all()

        counts: dict[str, int] = {}
        for row in refs:
            counts[row["resource_code"]] = counts.get(row["resource_code"], 0) + 1

        wb = Workbook()
        ws = wb.active
        ws.title = "專案資源"
        ws.append(RESOURCE_HEADERS)
        for row in resource_rows:
            ws.append([row["code"], row["name"], row["unit"], float(row["unit_price"]), counts.get(row["code"], 0)])
        ws2 = wb.create_sheet("引用工項")
        ws2.append(REFERENCE_HEADERS)
        for row in refs:
            ws2.append([row["resource_code"], row["item_no"], row["name"], float(row["quantity"]),
                        float(row["unit_price"]), float(row["amount"])])

        for sheet in (ws, ws2):
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for column, width in {"A":18, "B":36, "C":12, "D":16, "E":16, "F":16}.items():
                sheet.column_dimensions[column].width = width

        for cell in ws["D"][1:]: cell.number_format = self._fmt(policy["analysis_price_scale"])
        for cell in ws2["D"][1:]: cell.number_format = self._fmt(policy["main_quantity_scale"])
        for cell in ws2["E"][1:]: cell.number_format = self._fmt(policy["main_price_scale"])
        for cell in ws2["F"][1:]: cell.number_format = self._fmt(policy["main_amount_scale"])
        wb.properties.title = f"PCCES 專案資源 - {project_code}"
        wb.properties.subject = "Legacy FormBudgetRes compatible export"
        stream = BytesIO(); wb.save(stream)
        return stream.getvalue()


def build_mrs_excel_export_blueprint(service: MRSExcelExportService, resolve_user_id):
    bp = Blueprint("mrs_excel_export", __name__, url_prefix="/api/mrs")

    @bp.get("/projects/<project_code>/export.xlsx")
    def export_project(project_code: str):
        if resolve_user_id() is None:
            return jsonify({"code": "UNAUTHORIZED"}), 401
        payload = service.export_project(project_code)
        return Response(payload, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="mrs-{project_code}.xlsx"'})

    return bp
