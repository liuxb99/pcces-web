import unittest
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine

from api.budget_decimal import BudgetDecimalService
from api.mrs_excel_export import MRSExcelExportService, REFERENCE_HEADERS, RESOURCE_HEADERS
from api.mrs_precision_policy import MRSPrecisionPolicyService
from api.resource_budget_lineage import ResourceBudgetLineageService
from api.resource_decimal import ResourceDecimalService


class MRSExcelExportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
        self.budget = BudgetDecimalService(self.engine); self.budget.create_schema()
        self.resources = ResourceDecimalService(self.engine); self.resources.create_schema()
        self.links = ResourceBudgetLineageService(self.engine)
        self.precision = MRSPrecisionPolicyService(self.engine)
        self.budget.save("I1", {"project_code":"P1","item_no":"001","name":"混凝土工項","kind":"L",
            "quantity":"3.4567","unit_price":"12.3456","quantity_scale":4,"price_scale":4,"amount_scale":2,"row_version":0})
        self.resources.save_resource("R1", {"code":"M00001","name":"水泥","unit":"KG","unit_price":"7.8912","price_scale":4,"row_version":0})
        self.links.link("P1", "R1", "I1")

    def test_two_grid_workbook_and_legacy_headers(self):
        payload = MRSExcelExportService(self.engine).export_project("P1")
        wb = load_workbook(BytesIO(payload))
        self.assertEqual(["專案資源", "引用工項"], wb.sheetnames)
        self.assertEqual(RESOURCE_HEADERS, [c.value for c in wb["專案資源"][1]])
        self.assertEqual(REFERENCE_HEADERS, [c.value for c in wb["引用工項"][1]])
        self.assertEqual("M00001", wb["專案資源"]["A2"].value)
        self.assertEqual("001", wb["引用工項"]["B2"].value)
        self.assertEqual(1, wb["專案資源"]["E2"].value)

    def test_project_precision_controls_number_formats(self):
        current = self.precision.get("P1")
        self.precision.save("P1", {"main_quantity_scale":3,"main_price_scale":2,"main_amount_scale":1,
            "analysis_quantity_scale":5,"analysis_price_scale":4,"analysis_amount_scale":3,
            "row_version":current["row_version"]}, "u1")
        wb = load_workbook(BytesIO(MRSExcelExportService(self.engine).export_project("P1")))
        self.assertEqual("0.0000", wb["專案資源"]["D2"].number_format)
        self.assertEqual("0.000", wb["引用工項"]["D2"].number_format)
        self.assertEqual("0.00", wb["引用工項"]["E2"].number_format)
        self.assertEqual("0.0", wb["引用工項"]["F2"].number_format)

    def test_empty_project_still_exports_stable_schema(self):
        wb = load_workbook(BytesIO(MRSExcelExportService(self.engine).export_project("EMPTY")))
        self.assertEqual(1, wb["專案資源"].max_row)
        self.assertEqual(1, wb["引用工項"].max_row)
        self.assertEqual("A2", wb["專案資源"].freeze_panes)


if __name__ == "__main__":
    unittest.main()
