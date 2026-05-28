"""PCCES 網頁版 — Flask 主程式"""

import os
import json
import tempfile
from datetime import datetime, timedelta, timezone
from functools import wraps
import secrets
from hashlib import pbkdf2_hmac

import jwt
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from sqlalchemy import create_engine, func, select, delete, text
from sqlalchemy.orm import Session, sessionmaker

from api.models import Base, User, Project, BudgetItem, Resource, ResourceBreakdownItem, BudgetItemKind, UserRole, Invoice, InvoiceItem
from api.models import MrsBaseCategory, MrsBaseItem, MrsBaseBreakdownItem, MrsBaseBookmark
from api.models import Contract, ContractItem, ContractIssue, ContractIssueItem
from api.models import ContractSettlement, ContractSettlementItem
from api.models import ContractFinalAcceptance, ContractFinalAcceptanceItem
from api.models import SystemParameter, CodeTable, CodeItem, Organization, FeatureFlag
from api.version import APP_NAME, APP_VERSION, BUILD_DATE, REPO_URL, RELEASE_NOTES_URL, CHANGELOG, DEPENDENCIES

# ─── 設定（支援環境變數覆蓋） ───
import os
SECRET_KEY = os.environ.get("PCCES_SECRET_KEY", "pcces-web-secret-key-change-in-production")
_db_url_env = os.environ.get("PCCES_DATABASE_URL", "")
DATABASE_URL = _db_url_env if _db_url_env else f"sqlite:///{tempfile.gettempdir()}/pcces.db"
ALGORITHM = os.environ.get("PCCES_JWT_ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("PCCES_TOKEN_EXPIRE_MINUTES", "480"))
REPORT_DIR = os.environ.get("PCCES_REPORT_DIR", "reports")

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

# 資料庫
engine = create_engine(DATABASE_URL, echo=False)
SessionLocal = sessionmaker(bind=engine)


def init_db():
    Base.metadata.create_all(engine)
    _migrate_schema()


def _migrate_schema():
    """SQLite 遷移輔助：為既有表格補上新增的欄位（若不存在）"""
    from sqlalchemy import inspect
    inspector = inspect(engine)

    # ── resources 表新增欄位 ──
    cols = {c["name"] for c in inspector.get_columns("resources")}
    additions = {
        "is_analysis": "INTEGER DEFAULT 0",
        "labor_rate": "FLOAT DEFAULT 0",
        "material_rate": "FLOAT DEFAULT 0",
        "equipment_rate": "FLOAT DEFAULT 0",
        "misc_rate": "FLOAT DEFAULT 0",
    }
    for col, type_def in additions.items():
        if col not in cols:
            with engine.connect() as conn:
                conn.execute(f"ALTER TABLE resources ADD COLUMN {col} {type_def}")
                conn.commit()

    # ── resource_breakdown_items 表（若不存在則 create_all 已處理） ──


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ─── JWT 輔助函數 ───
def get_password_hash(password: str) -> str:
    """使用 PBKDF2-SHA256 加鹽雜湊密碼"""
    salt = secrets.token_hex(16)
    key = pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100000).hex()
    return f"{salt}${key}"


def verify_password(plain: str, stored: str) -> bool:
    """驗證密碼"""
    try:
        salt, key = stored.split("$")
        check = pbkdf2_hmac("sha256", plain.encode(), salt.encode(), 100000).hex()
        return check == key
    except (ValueError, AttributeError):
        return False


def create_token(user_id: int, username: str) -> str:
    payload = {
        "sub": str(user_id),
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.PyJWTError:
        return None


def require_auth(f):
    """免登入模式：有 token 就解析使用者，無 token 就用預設 guest"""
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            payload = decode_token(auth[7:])
            if payload:
                kwargs["user_id"] = int(payload["sub"])
                return f(*args, **kwargs)
        # 無 token 或 token 無效 → 以 guest 身份操作
        kwargs["user_id"] = 1
        return f(*args, **kwargs)
    return decorated


def require_admin(f):
    """admin 角色專用裝飾器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"detail": "未授權"}), 401
        payload = decode_token(auth[7:])
        if not payload:
            return jsonify({"detail": "Token 無效"}), 401
        user_id = int(payload["sub"])
        db = next(get_db())
        try:
            user = db.query(User).filter(User.id == user_id).first()
            if not user or user.role != UserRole.ADMIN.value:
                return jsonify({"detail": "需要管理員權限"}), 403
            kwargs["user_id"] = user_id
            return f(*args, **kwargs)
        finally:
            db.close()
    return decorated


# ─── 工具函數 ───
def round_value(value: float, decimals: int = 2) -> float:
    factor = 10 ** decimals
    if value < 0:
        return float(int(value * factor - 0.5)) / factor
    return float(int(value * factor + 0.5)) / factor


def model_to_dict(obj, exclude=None):
    """將 SQLAlchemy 模型轉為 dict"""
    if exclude is None:
        exclude = set()
    result = {}
    for col in obj.__table__.columns:
        if col.name not in exclude:
            val = getattr(obj, col.name)
            if isinstance(val, datetime):
                val = val.isoformat()
            result[col.name] = val
    return result


# ═══════════════════════════════════════════════
# 認證 API
# ═══════════════════════════════════════════════

@app.route("/api/health")
def health_check():
    return jsonify({"status": "ok", "version": "1.0.0"})


@app.route("/api/auth/register", methods=["POST"])
def register():
    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供資料"}), 400

    db = next(get_db())
    try:
        existing = db.query(User).filter(User.username == data["username"]).first()
        if existing:
            return jsonify({"detail": "帳號已存在"}), 400

        user = User(
            username=data["username"],
            password_hash=get_password_hash(data["password"]),
            display_name=data.get("display_name", data["username"]),
            email=data.get("email"),
            company=data.get("company"),
            department=data.get("department"),
        )
        db.add(user)
        db.commit()
        db.refresh(user)

        token = create_token(user.id, user.username)
        return jsonify({
            "access_token": token,
            "token_type": "bearer",
            "user": model_to_dict(user, exclude={"password_hash"}),
        }), 201
    finally:
        db.close()


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供帳號密碼"}), 400

    db = next(get_db())
    try:
        user = db.query(User).filter(User.username == data["username"]).first()
        if not user or not verify_password(data["password"], user.password_hash):
            return jsonify({"detail": "帳號或密碼錯誤"}), 401
        if not user.is_active:
            return jsonify({"detail": "帳號已停用"}), 403

        token = create_token(user.id, user.username)
        return jsonify({
            "access_token": token,
            "token_type": "bearer",
            "user": model_to_dict(user, exclude={"password_hash"}),
        })
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 專案 API
# ═══════════════════════════════════════════════

@app.route("/api/projects/stats")
@require_auth
def dashboard_stats(user_id):
    db = next(get_db())
    try:
        # 資料隔離：一般使用者只看自己的統計
        user = db.query(User).filter(User.id == user_id).first()
        is_admin = user and user.role == UserRole.ADMIN.value

        project_query = db.query(Project)
        if not is_admin:
            project_query = project_query.filter(Project.owner_id == user_id)

        total_projects = project_query.count() or 0
        active_projects = project_query.filter(Project.status == "active").count() or 0

        user_project_ids = [p.id for p in project_query.all()]
        if user_project_ids:
            total_items = db.query(func.count(BudgetItem.id)).filter(
                BudgetItem.project_id.in_(user_project_ids)
            ).scalar() or 0
            total_amount = db.query(func.coalesce(func.sum(BudgetItem.amount), 0)).filter(
                BudgetItem.project_id.in_(user_project_ids)
            ).scalar() or 0.0
            total_resources = db.query(func.count(Resource.id)).filter(
                Resource.project_id.in_(user_project_ids)
            ).scalar() or 0
        else:
            total_items = total_amount = total_resources = 0

        recent_query = db.query(Project).order_by(Project.updated_at.desc())
        if not is_admin:
            recent_query = recent_query.filter(Project.owner_id == user_id)
        recent = recent_query.limit(5).all()

        return jsonify({
            "total_projects": total_projects,
            "active_projects": active_projects,
            "total_budget_items": total_items,
            "total_budget_amount": float(total_amount),
            "total_resources": total_resources,
            "recent_projects": [model_to_dict(p) for p in recent],
        })
    finally:
        db.close()


@app.route("/api/projects/", methods=["GET"])
@require_auth
def list_projects(user_id):
    db = next(get_db())
    try:
        # 資料隔離：一般使用者只看自己專案，管理員可看全部
        user = db.query(User).filter(User.id == user_id).first()
        is_admin = user and user.role == UserRole.ADMIN.value
        query = db.query(Project).order_by(Project.updated_at.desc())
        if not is_admin:
            query = query.filter(Project.owner_id == user_id)
        projects = query.all()
        result = []
        for p in projects:
            d = model_to_dict(p)
            # 計算預算總額與項目數
            items = db.query(BudgetItem).filter(
                BudgetItem.project_id == p.id,
                BudgetItem.parent_id.is_(None)
            ).all()
            d["budget_total"] = sum((i.amount or 0) for i in items)
            d["item_count"] = db.query(func.count(BudgetItem.id)).filter(
                BudgetItem.project_id == p.id
            ).scalar() or 0
            result.append(d)
        return jsonify(result)
    finally:
        db.close()


@app.route("/api/projects/", methods=["POST"])
@require_auth
def create_project(user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        project = Project(
            code=data["code"],
            name=data["name"],
            name_en=data.get("name_en"),
            location=data.get("location"),
            account_code=data.get("account_code"),
            description=data.get("description"),
            scope=data.get("scope", 1.0),
            scope_unit=data.get("scope_unit", "式"),
            owner_id=user_id,
        )
        db.add(project)
        db.commit()
        db.refresh(project)
        return jsonify(model_to_dict(project)), 201
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>", methods=["GET"])
@require_auth
def get_project(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        d = model_to_dict(proj)
        items = db.query(BudgetItem).filter(
            BudgetItem.project_id == project_id,
            BudgetItem.parent_id.is_(None)
        ).all()
        d["budget_total"] = sum((i.amount or 0) for i in items)
        d["item_count"] = db.query(func.count(BudgetItem.id)).filter(
            BudgetItem.project_id == project_id
        ).scalar() or 0
        return jsonify(d)
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>", methods=["PUT"])
@require_auth
def update_project(project_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            return jsonify({"detail": "專案不存在"}), 404
        # 所有權檢查
        user = db.query(User).filter(User.id == user_id).first()
        is_admin = user and user.role == UserRole.ADMIN.value
        if not is_admin and project.owner_id != user_id:
            return jsonify({"detail": "無權限修改此專案"}), 403
        for key in ("name", "name_en", "location", "account_code", "description", "scope", "scope_unit", "status"):
            if key in data and data[key] is not None:
                setattr(project, key, data[key])
        db.commit()
        db.refresh(project)
        return jsonify(model_to_dict(project))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>", methods=["DELETE"])
@require_auth
def delete_project(project_id, user_id):
    db = next(get_db())
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            return jsonify({"detail": "專案不存在"}), 404
        # 所有權檢查
        user = db.query(User).filter(User.id == user_id).first()
        is_admin = user and user.role == UserRole.ADMIN.value
        if not is_admin and project.owner_id != user_id:
            return jsonify({"detail": "無權限刪除專案"}), 403
        db.delete(project)
        db.commit()
        return jsonify({"message": "專案已刪除"})
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 預算項目 API（核心）
# ═══════════════════════════════════════════════

def _get_children_count(db: Session, parent_id: int) -> int:
    return db.query(func.count(BudgetItem.id)).filter(
        BudgetItem.parent_id == parent_id
    ).scalar() or 0


def _calc_amount(item: BudgetItem) -> float:
    return round_value((item.quantity or 0) * (item.unit_price or 0), item.decimal_amount)


def _recalc_children(db: Session, parent_id: int) -> float:
    """遞迴計算子項金額"""
    children = db.query(BudgetItem).filter(BudgetItem.parent_id == parent_id).all()
    total = 0.0
    for child in children:
        if child.kind in (BudgetItemKind.B, BudgetItemKind.Z):
            child.amount = _recalc_children(db, child.id)
        else:
            child.amount = _calc_amount(child)
        db.flush()
        total += child.amount or 0
    return round_value(total, 2)


def _check_project_access(db: Session, project_id: int, user_id: int):
    """檢查使用者是否有權限存取專案（回傳 (project, error_response)）"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        return None, (jsonify({"detail": "專案不存在"}), 404)
    user = db.query(User).filter(User.id == user_id).first()
    is_admin = user and user.role == UserRole.ADMIN.value
    if not is_admin and project.owner_id != user_id:
        return None, (jsonify({"detail": "無權限操作此專案"}), 403)
    return project, None


def _build_tree_dict(db: Session, project_id: int, parent_id=None):
    """遞迴建立預算樹狀結構（含 children 巢狀資料）"""
    query = db.query(BudgetItem).filter(BudgetItem.project_id == project_id)
    if parent_id is None:
        query = query.filter(BudgetItem.parent_id.is_(None))
    else:
        query = query.filter(BudgetItem.parent_id == parent_id)
    query = query.order_by(BudgetItem.sort_order, BudgetItem.id)
    items = query.all()
    result = []
    for item in items:
        d = model_to_dict(item)
        d["children"] = _build_tree_dict(db, project_id, item.id)
        result.append(d)
    return result


@app.route("/api/projects/<int:project_id>/budget/tree", methods=["GET"])
@require_auth
def get_budget_tree(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        tree = _build_tree_dict(db, project_id)
        return jsonify(tree)
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/budget/", methods=["GET"])
@require_auth
def get_budget_list(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        items = db.query(BudgetItem).filter(
            BudgetItem.project_id == project_id
        ).order_by(BudgetItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/budget/", methods=["POST"])
@require_auth
def create_budget_item(project_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        item = BudgetItem(
            project_id=project_id,
            parent_id=data.get("parent_id"),
            item_no=data.get("item_no"),
            print_no=data.get("print_no"),
            c_name=data.get("c_name"),
            e_name=data.get("e_name"),
            c_unit=data.get("c_unit"),
            e_unit=data.get("e_unit"),
            quantity=data.get("quantity", 0),
            unit_price=data.get("unit_price", 0),
            kind=BudgetItemKind(data.get("kind", "B")),
            formula=data.get("formula"),
            memo=data.get("memo"),
            sort_order=data.get("sort_order"),
            is_fixed_price=data.get("is_fixed_price", False),
            decimal_qty=data.get("decimal_qty", 2),
            decimal_price=data.get("decimal_price", 2),
            decimal_amount=data.get("decimal_amount", 2),
        )
        # B/Z 類型金額不計算（需透過 recalc 加總子項）
        kind_val = item.kind.value if hasattr(item.kind, 'value') else str(item.kind)
        if kind_val not in ('B', 'Z'):
            item.amount = _calc_amount(item)
        db.add(item)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/budget/<int:item_id>", methods=["PUT"])
@require_auth
def update_budget_item(project_id, item_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        item = db.query(BudgetItem).filter(
            BudgetItem.id == item_id,
            BudgetItem.project_id == project_id
        ).first()
        if not item:
            return jsonify({"detail": "預算項目不存在"}), 404

        for key in ("parent_id", "item_no", "print_no", "c_name", "e_name",
                     "c_unit", "e_unit", "formula", "memo", "sort_order",
                     "is_fixed_price", "is_locked", "decimal_qty", "decimal_price",
                     "decimal_amount", "is_green_item"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])

        if "quantity" in data:
            item.quantity = data["quantity"]
        if "unit_price" in data:
            item.unit_price = data["unit_price"]
        if "kind" in data:
            item.kind = BudgetItemKind(data["kind"])

        # B/Z 類型金額不計算（需透過 recalc 加總子項）
        kind_val = item.kind.value if hasattr(item.kind, 'value') else str(item.kind)
        if kind_val not in ('B', 'Z'):
            item.amount = _calc_amount(item)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/budget/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_budget_item(project_id, item_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        item = db.query(BudgetItem).filter(
            BudgetItem.id == item_id,
            BudgetItem.project_id == project_id
        ).first()
        if not item:
            return jsonify({"detail": "預算項目不存在"}), 404
        # 先遞迴刪除所有子孫項目，再刪除本身
        _delete_item_children(db, item_id)
        db.delete(item)
        db.commit()
        return jsonify({"message": "預算項目已刪除"})
    finally:
        db.close()


def _delete_item_children(db: Session, parent_id: int):
    """遞迴刪除子項目"""
    children = db.query(BudgetItem).filter(BudgetItem.parent_id == parent_id).all()
    for child in children:
        _delete_item_children(db, child.id)
        db.delete(child)


@app.route("/api/projects/<int:project_id>/budget/<int:item_id>/move", methods=["POST"])
@require_auth
def move_budget_item(project_id, item_id, user_id):
    # 支援 query param 與 JSON body 兩種格式
    new_parent_id = request.args.get("new_parent_id")
    if new_parent_id is None:
        data = request.get_json() or {}
        new_parent_id = data.get("new_parent_id")
    if new_parent_id is not None:
        new_parent_id = int(new_parent_id) if str(new_parent_id) != 'null' else None
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        item = db.query(BudgetItem).filter(
            BudgetItem.id == item_id,
            BudgetItem.project_id == project_id
        ).first()
        if not item:
            return jsonify({"detail": "預算項目不存在"}), 404
        item.parent_id = new_parent_id
        db.commit()
        return jsonify({"message": "預算項目已移動"})
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/budget/recalc", methods=["POST"])
@require_auth
def recalc_budget(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        root_items = db.query(BudgetItem).filter(
            BudgetItem.project_id == project_id,
            BudgetItem.parent_id.is_(None)
        ).all()
        for item in root_items:
            if item.kind == BudgetItemKind.B:
                item.amount = _recalc_children(db, item.id)
            else:
                item.amount = _calc_amount(item)
        db.commit()
        return jsonify({"message": "預算重新計算完成"})
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 資源 API
# ═══════════════════════════════════════════════

@app.route("/api/projects/<int:project_id>/resources/", methods=["GET"])
@require_auth
def list_resources(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resources = db.query(Resource).filter(Resource.project_id == project_id).all()
        return jsonify([model_to_dict(r) for r in resources])
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/", methods=["POST"])
@require_auth
def create_resource(project_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resource = Resource(
            project_id=project_id,
            code=data["code"],
            c_name=data["c_name"],
            e_name=data.get("e_name"),
            c_unit=data.get("c_unit", "式"),
            e_unit=data.get("e_unit"),
            unit_price=data.get("unit_price", 0),
            category=data.get("category", "material"),
            remark=data.get("remark"),
        )
        db.add(resource)
        db.commit()
        db.refresh(resource)
        return jsonify(model_to_dict(resource)), 201
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/<int:resource_id>/price", methods=["PUT"])
@require_auth
def update_resource_price(project_id, resource_id, user_id):
    # 支援 query param 與 JSON body 兩種格式
    unit_price = request.args.get("unit_price")
    if unit_price is None:
        data = request.get_json() or {}
        unit_price = data.get("unit_price")
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resource = db.query(Resource).filter(
            Resource.id == resource_id,
            Resource.project_id == project_id
        ).first()
        if not resource:
            return jsonify({"detail": "資源不存在"}), 404
        if unit_price is not None:
            resource.unit_price = float(unit_price)
        db.commit()
        db.refresh(resource)
        return jsonify(model_to_dict(resource))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/<int:resource_id>", methods=["PUT"])
@require_auth
def update_resource(project_id, resource_id, user_id):
    """更新資源欄位（含單價分析相關欄位）"""
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resource = db.query(Resource).filter(
            Resource.id == resource_id,
            Resource.project_id == project_id
        ).first()
        if not resource:
            return jsonify({"detail": "資源不存在"}), 404

        # 可更新的欄位
        for key in ("code", "c_name", "e_name", "c_unit", "e_unit", "category",
                     "remark", "is_public", "is_analysis"):
            if key in data and data[key] is not None:
                setattr(resource, key, data[key])

        if "unit_price" in data and data["unit_price"] is not None:
            resource.unit_price = float(data["unit_price"])
        if "labor_rate" in data and data["labor_rate"] is not None:
            resource.labor_rate = float(data["labor_rate"])
        if "material_rate" in data and data["material_rate"] is not None:
            resource.material_rate = float(data["material_rate"])
        if "equipment_rate" in data and data["equipment_rate"] is not None:
            resource.equipment_rate = float(data["equipment_rate"])
        if "misc_rate" in data and data["misc_rate"] is not None:
            resource.misc_rate = float(data["misc_rate"])

        db.commit()
        db.refresh(resource)
        return jsonify(model_to_dict(resource))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/analysis", methods=["GET"])
@require_auth
def list_resources_analysis(project_id, user_id):
    """列出啟用單價分析的資源（含各細項）"""
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resources = db.query(Resource).filter(
            Resource.project_id == project_id,
            Resource.is_analysis == True
        ).all()
        result = []
        for r in resources:
            d = model_to_dict(r)
            items = db.query(ResourceBreakdownItem).filter(
                ResourceBreakdownItem.resource_id == r.id
            ).all()
            d["breakdown_items"] = [model_to_dict(i) for i in items]
            # 計算細項總金額
            d["breakdown_total"] = sum((i.amount or 0) for i in items)
            result.append(d)
        return jsonify(result)
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/<int:resource_id>/breakdown", methods=["GET"])
@require_auth
def list_resource_breakdown(project_id, resource_id, user_id):
    """取得資源單價分析細項列表"""
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resource = db.query(Resource).filter(
            Resource.id == resource_id,
            Resource.project_id == project_id
        ).first()
        if not resource:
            return jsonify({"detail": "資源不存在"}), 404
        items = db.query(ResourceBreakdownItem).filter(
            ResourceBreakdownItem.resource_id == resource_id
        ).order_by(ResourceBreakdownItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/<int:resource_id>/breakdown", methods=["POST"])
@require_auth
def create_resource_breakdown(project_id, resource_id, user_id):
    """新增資源單價分析細項"""
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resource = db.query(Resource).filter(
            Resource.id == resource_id,
            Resource.project_id == project_id
        ).first()
        if not resource:
            return jsonify({"detail": "資源不存在"}), 404

        qty = float(data.get("quantity", 0))
        up = float(data.get("unit_price", 0))
        item = ResourceBreakdownItem(
            resource_id=resource_id,
            code=data.get("code", ""),
            c_name=data.get("c_name", ""),
            c_unit=data.get("c_unit", ""),
            quantity=qty,
            unit_price=up,
            amount=round(qty * up, 2),
            remark=data.get("remark"),
        )
        db.add(item)
        db.flush()

        # 若資源啟用單價分析，計算加總金額寫回 unit_price
        total = db.query(func.coalesce(func.sum(ResourceBreakdownItem.amount), 0)).filter(
            ResourceBreakdownItem.resource_id == resource_id
        ).scalar() or 0.0
        if resource.is_analysis:
            resource.unit_price = round(float(total), 2)

        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/<int:resource_id>/breakdown/<int:breakdown_id>", methods=["DELETE"])
@require_auth
def delete_resource_breakdown(project_id, resource_id, breakdown_id, user_id):
    """刪除資源單價分析細項，並重新計算總金額"""
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resource = db.query(Resource).filter(
            Resource.id == resource_id,
            Resource.project_id == project_id
        ).first()
        if not resource:
            return jsonify({"detail": "資源不存在"}), 404
        item = db.query(ResourceBreakdownItem).filter(
            ResourceBreakdownItem.id == breakdown_id,
            ResourceBreakdownItem.resource_id == resource_id
        ).first()
        if not item:
            return jsonify({"detail": "細項不存在"}), 404

        db.delete(item)
        db.flush()

        # 重新計算單價 = 細項加總
        total = db.query(func.coalesce(func.sum(ResourceBreakdownItem.amount), 0)).filter(
            ResourceBreakdownItem.resource_id == resource_id
        ).scalar() or 0.0
        if resource.is_analysis:
            resource.unit_price = round(float(total), 2)

        db.commit()
        return jsonify({"message": "細項已刪除"})
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/resources/analysis/recalc", methods=["POST"])
@require_auth
def recalc_resource_analysis(project_id, user_id):
    """重新計算所有啟用分析的資源單價（細項加總）"""
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        resources = db.query(Resource).filter(
            Resource.project_id == project_id,
            Resource.is_analysis == True
        ).all()
        for r in resources:
            total = db.query(func.coalesce(func.sum(ResourceBreakdownItem.amount), 0)).filter(
                ResourceBreakdownItem.resource_id == r.id
            ).scalar() or 0.0
            r.unit_price = round(float(total), 2)
        db.commit()
        return jsonify({"message": "資源單價分析重新計算完成"})
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 報表 API
# ═══════════════════════════════════════════════

@app.route("/api/projects/<int:project_id>/reports/summary")
@require_auth
def get_summary_report(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        items = db.query(BudgetItem).filter(BudgetItem.project_id == project_id).all()
        root_items = [i for i in items if i.parent_id is None]
        total_amount = sum(i.amount or 0 for i in root_items)

        category_totals = {}
        for i in root_items:
            k = i.kind.value if hasattr(i.kind, 'value') else str(i.kind)
            category_totals[k] = category_totals.get(k, 0) + (i.amount or 0)

        return jsonify({
            "project_id": project_id,
            "total_amount": total_amount,
            "item_count": len(items),
            "root_count": len(root_items),
            "category_totals": category_totals,
            "generated_at": datetime.now().isoformat(),
        })
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/reports/excel")
@require_auth
def export_excel(project_id, user_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            return jsonify({"detail": "專案不存在"}), 404

        items = db.query(BudgetItem).filter(
            BudgetItem.project_id == project_id
        ).order_by(BudgetItem.sort_order, BudgetItem.id).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "預算總表"

        title_font = Font(name="微軟正黑體", size=16, bold=True)
        header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="微軟正黑體", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        ws.merge_cells('A1:H1')
        ws['A1'] = f"公共工程經費估算表 — {project.name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:H2')
        ws['A2'] = f"專案編號: {project.code} | {datetime.now().strftime('%Y/%m/%d')}"
        ws['A2'].font = Font(name="微軟正黑體", size=10, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ["項次", "項目編號", "項目名稱", "單位", "數量", "單價", "複價", "備註"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        row = 5
        item_map = {i.id: i for i in items}

        def write_item(item, level=0):
            nonlocal row
            ws.cell(row=row, column=1, value=item.print_no or "").font = normal_font
            ws.cell(row=row, column=2, value=item.item_no or "").font = normal_font
            ws.cell(row=row, column=3, value=("  " * level) + (item.c_name or "")).font = normal_font
            ws.cell(row=row, column=4, value=item.c_unit or "").font = normal_font
            ws.cell(row=row, column=5, value=item.quantity or "").font = normal_font
            ws.cell(row=row, column=6, value=item.unit_price or "").font = normal_font
            ws.cell(row=row, column=7, value=item.amount or "").font = normal_font
            ws.cell(row=row, column=8, value=item.memo or "").font = normal_font
            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).alignment = Alignment(horizontal='center' if c <= 2 else 'left')
            row += 1

            # 子項
            children = [i for i in items if i.parent_id == item.id]
            for child in children:
                write_item(child, level + 1)

        # 只寫根節點
        root_items = [i for i in items if i.parent_id is None]
        for item in root_items:
            write_item(item)

        # 合計
        total = sum(i.amount or 0 for i in root_items)
        ws.cell(row=row, column=6, value="總計").font = Font(name="微軟正黑體", size=11, bold=True)
        ws.cell(row=row, column=7, value=total).font = Font(name="微軟正黑體", size=11, bold=True)
        ws.cell(row=row, column=7).number_format = '#,##0'
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 20

        os.makedirs(REPORT_DIR, exist_ok=True)
        filepath = os.path.join(REPORT_DIR, f"budget_{project_id}.xlsx")
        wb.save(filepath)

        return send_file(
            filepath,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"PCCES_{project.code}_預算表.xlsx",
        )
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 計價管理 API（Invoice）
# ═══════════════════════════════════════════════

def _calc_invoice_item_amounts(item: InvoiceItem):
    """計算 InvoiceItem 的金額與完成率"""
    item.total_completed_qty = (item.prev_completed_qty or 0) + (item.this_completed_qty or 0)
    item.this_amount = round((item.this_completed_qty or 0) * (item.unit_price or 0), 2)
    item.cumulative_amount = round((item.total_completed_qty or 0) * (item.unit_price or 0), 2)
    item.remain_qty = max(0, (item.contract_qty or 0) - item.total_completed_qty)
    if item.contract_qty and item.contract_qty > 0:
        item.progress_rate = round((item.total_completed_qty / item.contract_qty) * 100, 2)
    else:
        item.progress_rate = 0


def _recalc_invoice(db: Session, invoice_id: int):
    """重算 Invoice 所有明細金額並更新主檔累計值"""
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        return
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_id).all()
    total_this = 0.0
    total_cum = 0.0
    for item in items:
        _calc_invoice_item_amounts(item)
        total_this += item.this_amount or 0
        total_cum += item.cumulative_amount or 0
        db.flush()
    inv.total_amount = round(total_this, 2)
    inv.cumulative_amount = round(total_cum, 2)
    # progress_rate = 與所有合約數量相比的加權進度（以金額比）
    total_contract_amount = db.query(func.coalesce(func.sum(InvoiceItem.contract_qty * InvoiceItem.unit_price), 0)).filter(
        InvoiceItem.invoice_id == invoice_id
    ).scalar() or 0.0
    if total_contract_amount > 0:
        inv.progress_rate = round((inv.cumulative_amount / total_contract_amount) * 100, 2)
    else:
        inv.progress_rate = 0
    db.flush()


# ─── Invoices CRUD ───

@app.route("/api/projects/<int:project_id>/invoices/", methods=["GET"])
@require_auth
def list_invoices(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        invoices = db.query(Invoice).filter(
            Invoice.project_id == project_id
        ).order_by(Invoice.period_no.desc(), Invoice.created_at.desc()).all()
        return jsonify([model_to_dict(inv) for inv in invoices])
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/invoices/", methods=["POST"])
@require_auth
def create_invoice(project_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        # 自動產生期別：以最新期別 +1
        last = db.query(func.max(Invoice.period_no)).filter(
            Invoice.project_id == project_id
        ).scalar() or 0
        period_no = last + 1

        inv = Invoice(
            project_id=project_id,
            period_no=period_no,
            invoice_no=data.get("invoice_no", f"INV-{period_no:04d}"),
            c_name=data.get("c_name", f"第{period_no}期計價"),
            status="draft",
            description=data.get("description"),
            invoice_date=data.get("invoice_date"),
            remark=data.get("remark"),
            created_by=user_id,
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        return jsonify(model_to_dict(inv)), 201
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/invoices/<int:invoice_id>", methods=["GET"])
@require_auth
def get_invoice(project_id, invoice_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        inv = db.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.project_id == project_id
        ).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        d = model_to_dict(inv)
        # 附帶明細數量
        d["item_count"] = db.query(func.count(InvoiceItem.id)).filter(
            InvoiceItem.invoice_id == invoice_id
        ).scalar() or 0
        return jsonify(d)
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/invoices/<int:invoice_id>", methods=["PUT"])
@require_auth
def update_invoice(project_id, invoice_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        inv = db.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.project_id == project_id
        ).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        for key in ("invoice_no", "c_name", "description", "invoice_date", "remark"):
            if key in data and data[key] is not None:
                setattr(inv, key, data[key])
        db.commit()
        db.refresh(inv)
        return jsonify(model_to_dict(inv))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/invoices/<int:invoice_id>", methods=["DELETE"])
@require_auth
def delete_invoice(project_id, invoice_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        inv = db.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.project_id == project_id
        ).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可刪除"}), 400
        # 明細會因 CASCADE 自動刪除
        db.delete(inv)
        db.commit()
        return jsonify({"message": "計價單已刪除"})
    finally:
        db.close()


# ─── Invoice Items ───

@app.route("/api/invoices/<int:invoice_id>/items/", methods=["GET"])
@require_auth
def list_invoice_items(invoice_id, user_id):
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        items = db.query(InvoiceItem).filter(
            InvoiceItem.invoice_id == invoice_id
        ).order_by(InvoiceItem.sort_order, InvoiceItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/items/", methods=["POST"])
@require_auth
def create_invoice_item(invoice_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        contract_qty = float(data.get("contract_qty", 0))
        unit_price = float(data.get("unit_price", 0))
        prev_completed = float(data.get("prev_completed_qty", 0))
        this_completed = float(data.get("this_completed_qty", 0))

        item = InvoiceItem(
            invoice_id=invoice_id,
            budget_item_id=data.get("budget_item_id"),
            item_no=data.get("item_no"),
            print_no=data.get("print_no"),
            c_name=data.get("c_name"),
            c_unit=data.get("c_unit"),
            contract_qty=contract_qty,
            unit_price=unit_price,
            prev_completed_qty=prev_completed,
            this_completed_qty=this_completed,
            sort_order=data.get("sort_order"),
            remark=data.get("remark"),
        )
        _calc_invoice_item_amounts(item)
        db.add(item)
        db.flush()
        _recalc_invoice(db, invoice_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/items/batch", methods=["POST"])
@require_auth
def batch_create_invoice_items(invoice_id, user_id):
    """批次從預算項目建立計價明細"""
    data = request.get_json() or {}
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        # 若傳入 budget_item_ids 則只匯入指定項目；否則匯入所有 W 類型項目
        item_ids = data.get("budget_item_ids")
        budget_query = db.query(BudgetItem).filter(
            BudgetItem.project_id == inv.project_id,
            BudgetItem.kind == BudgetItemKind.W.value,
        )
        if item_ids:
            budget_query = budget_query.filter(BudgetItem.id.in_(item_ids))
        budget_items = budget_query.all()

        # 取得當前 invoice 已有的 budget_item_id 集合（避免重複匯入）
        existing_ids = set()
        existing = db.query(InvoiceItem.budget_item_id).filter(
            InvoiceItem.invoice_id == invoice_id,
            InvoiceItem.budget_item_id.isnot(None),
        ).all()
        for (eid,) in existing:
            existing_ids.add(eid)

        created = []
        for bi in budget_items:
            if bi.id in existing_ids:
                continue
            item = InvoiceItem(
                invoice_id=invoice_id,
                budget_item_id=bi.id,
                item_no=bi.item_no,
                print_no=bi.print_no,
                c_name=bi.c_name,
                c_unit=bi.c_unit,
                contract_qty=bi.quantity or 0,
                unit_price=bi.unit_price or 0,
                prev_completed_qty=0,
                this_completed_qty=0,
                sort_order=bi.sort_order,
            )
            _calc_invoice_item_amounts(item)
            db.add(item)
            db.flush()
            created.append(model_to_dict(item))

        _recalc_invoice(db, invoice_id)
        db.commit()
        return jsonify({"created": created, "count": len(created)}), 201
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_invoice_item(invoice_id, item_id, user_id):
    """更新計價明細（完成數量等）"""
    data = request.get_json()
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        item = db.query(InvoiceItem).filter(
            InvoiceItem.id == item_id,
            InvoiceItem.invoice_id == invoice_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404

        for key in ("this_completed_qty", "prev_completed_qty", "contract_qty",
                     "unit_price", "remark", "sort_order", "c_name", "c_unit",
                     "item_no", "print_no"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])

        _calc_invoice_item_amounts(item)
        db.flush()
        _recalc_invoice(db, invoice_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/items/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_invoice_item(invoice_id, item_id, user_id):
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        item = db.query(InvoiceItem).filter(
            InvoiceItem.id == item_id,
            InvoiceItem.invoice_id == invoice_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404
        db.delete(item)
        db.flush()
        _recalc_invoice(db, invoice_id)
        db.commit()
        return jsonify({"message": "明細已刪除"})
    finally:
        db.close()


# ─── Invoice Operations ───

@app.route("/api/invoices/<int:invoice_id>/recalc", methods=["POST"])
@require_auth
def recalc_invoice(invoice_id, user_id):
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        _recalc_invoice(db, invoice_id)
        db.commit()
        return jsonify({"message": "計價金額重新計算完成"})
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/submit", methods=["POST"])
@require_auth
def submit_invoice(invoice_id, user_id):
    """提交審核"""
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        if inv.status != "draft":
            return jsonify({"detail": "僅草稿狀態可提交"}), 400
        # 確保至少有一筆明細
        item_count = db.query(func.count(InvoiceItem.id)).filter(
            InvoiceItem.invoice_id == invoice_id
        ).scalar() or 0
        if item_count == 0:
            return jsonify({"detail": "請先加入計價明細再提交"}), 400
        # 提交前重算
        _recalc_invoice(db, invoice_id)
        inv.status = "submitted"
        db.commit()
        db.refresh(inv)
        return jsonify(model_to_dict(inv))
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/approve", methods=["POST"])
@require_auth
def approve_invoice(invoice_id, user_id):
    """核准計價單"""
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err
        if inv.status != "submitted":
            return jsonify({"detail": "僅已提交狀態可核准"}), 400
        inv.status = "approved"
        db.commit()
        db.refresh(inv)
        return jsonify(model_to_dict(inv))
    finally:
        db.close()


# ─── Invoice Reports ───

@app.route("/api/invoices/<int:invoice_id>/report")
@require_auth
def invoice_report(invoice_id, user_id):
    """HTML 報表預覽"""
    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err

        items = db.query(InvoiceItem).filter(
            InvoiceItem.invoice_id == invoice_id
        ).order_by(InvoiceItem.sort_order, InvoiceItem.id).all()

        # 計算各項統計
        total_contract = sum((i.contract_qty or 0) * (i.unit_price or 0) for i in items)
        total_prev = sum((i.prev_completed_qty or 0) * (i.unit_price or 0) for i in items)
        total_this = inv.total_amount or 0
        total_cum = inv.cumulative_amount or 0
        total_remain = sum((i.remain_qty or 0) * (i.unit_price or 0) for i in items)

        rows_html = ""
        for idx, item in enumerate(items, 1):
            rows_html += f"""<tr>
                <td>{idx}</td>
                <td>{item.item_no or ''}</td>
                <td>{item.c_name or ''}</td>
                <td>{item.c_unit or ''}</td>
                <td style="text-align:right">{item.contract_qty or 0:,.2f}</td>
                <td style="text-align:right">{item.unit_price or 0:,.2f}</td>
                <td style="text-align:right">{item.prev_completed_qty or 0:,.2f}</td>
                <td style="text-align:right">{item.this_completed_qty or 0:,.2f}</td>
                <td style="text-align:right">{item.total_completed_qty or 0:,.2f}</td>
                <td style="text-align:right">{item.this_amount or 0:,.2f}</td>
                <td style="text-align:right">{item.cumulative_amount or 0:,.2f}</td>
                <td style="text-align:right">{item.remain_qty or 0:,.2f}</td>
                <td style="text-align:right">{item.progress_rate or 0:.1f}%</td>
                <td>{item.remark or ''}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<title>計價報表 - {inv.c_name}</title>
<style>
  body {{ font-family: 'Microsoft JhengHei', Arial, sans-serif; margin: 20px; }}
  h1 {{ font-size: 18px; text-align: center; margin-bottom: 5px; }}
  .subtitle {{ text-align: center; color: #666; font-size: 12px; margin-bottom: 20px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 11px; }}
  th {{ background: #4472C4; color: #fff; padding: 5px; border: 1px solid #999; text-align: center; }}
  td {{ padding: 4px 5px; border: 1px solid #999; }}
  .summary {{ margin-top: 15px; font-size: 13px; }}
  .summary td {{ padding: 6px 10px; }}
  .text-right {{ text-align: right; }}
</style>
</head>
<body>
  <h1>{proj.name} — 計價報表</h1>
  <div class="subtitle">
    計價單號：{inv.invoice_no or ''} | 期別：第{inv.period_no}期 | 狀態：{inv.status}
    | 日期：{inv.invoice_date or ''}
  </div>
  <table>
    <thead>
      <tr>
        <th>項次</th><th>編號</th><th>項目名稱</th><th>單位</th>
        <th>合約數量</th><th>單價</th><th>前期完成</th><th>本期完成</th>
        <th>累計完成</th><th>本期金額</th><th>累計金額</th>
        <th>剩餘數量</th><th>進度</th><th>備註</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
  <table class="summary">
    <tr><td><strong>合約總價：</strong>{total_contract:,.2f}</td>
        <td><strong>前期累計：</strong>{total_prev:,.2f}</td>
        <td><strong>本期計價：</strong>{total_this:,.2f}</td>
        <td><strong>累計計價：</strong>{total_cum:,.2f}</td>
        <td><strong>剩餘金額：</strong>{total_remain:,.2f}</td>
        <td><strong>完成進度：</strong>{inv.progress_rate:.1f}%</td></tr>
  </table>
  <p style="text-align:center; color:#999; font-size:10px; margin-top:30px;">
    產製時間：{datetime.now(timezone.utc).strftime('%Y/%m/%d %H:%M')}
  </p>
</body>
</html>"""
        return html, 200, {"Content-Type": "text/html; charset=utf-8"}
    finally:
        db.close()


@app.route("/api/invoices/<int:invoice_id>/export/excel")
@require_auth
def export_invoice_excel(invoice_id, user_id):
    """Excel 匯出"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    db = next(get_db())
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            return jsonify({"detail": "計價單不存在"}), 404
        proj, err = _check_project_access(db, inv.project_id, user_id)
        if err:
            return err

        items = db.query(InvoiceItem).filter(
            InvoiceItem.invoice_id == invoice_id
        ).order_by(InvoiceItem.sort_order, InvoiceItem.id).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "計價報表"

        title_font = Font(name="微軟正黑體", size=14, bold=True)
        header_font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
        normal_font = Font(name="微軟正黑體", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        ws.merge_cells('A1:N1')
        ws['A1'] = f"{proj.name} — 計價報表"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:N2')
        ws['A2'] = (f"計價單號：{inv.invoice_no or ''} | 期別：第{inv.period_no}期 | "
                    f"狀態：{inv.status} | 日期：{inv.invoice_date or ''}")
        ws['A2'].font = Font(name="微軟正黑體", size=9, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ["項次", "編號", "項目名稱", "單位", "合約數量", "單價",
                    "前期完成", "本期完成", "累計完成", "本期金額", "累計金額",
                    "剩餘數量", "進度(%)", "備註"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for idx, item in enumerate(items, 1):
            row = idx + 4
            ws.cell(row=row, column=1, value=idx).font = normal_font
            ws.cell(row=row, column=2, value=item.item_no or "").font = normal_font
            ws.cell(row=row, column=3, value=item.c_name or "").font = normal_font
            ws.cell(row=row, column=4, value=item.c_unit or "").font = normal_font
            ws.cell(row=row, column=5, value=item.contract_qty or 0).font = normal_font
            ws.cell(row=row, column=6, value=item.unit_price or 0).font = normal_font
            ws.cell(row=row, column=7, value=item.prev_completed_qty or 0).font = normal_font
            ws.cell(row=row, column=8, value=item.this_completed_qty or 0).font = normal_font
            ws.cell(row=row, column=9, value=item.total_completed_qty or 0).font = normal_font
            ws.cell(row=row, column=10, value=item.this_amount or 0).font = normal_font
            ws.cell(row=row, column=11, value=item.cumulative_amount or 0).font = normal_font
            ws.cell(row=row, column=12, value=item.remain_qty or 0).font = normal_font
            ws.cell(row=row, column=13, value=item.progress_rate or 0).font = normal_font
            ws.cell(row=row, column=14, value=item.remark or "").font = normal_font
            for c in range(1, 15):
                ws.cell(row=row, column=c).border = thin_border

        # 合計行
        total_row = len(items) + 5
        ws.cell(row=total_row, column=1, value="合計").font = Font(name="微軟正黑體", size=10, bold=True)
        ws.cell(row=total_row, column=5, value=sum(i.contract_qty or 0 for i in items))
        ws.cell(row=total_row, column=7, value=sum(i.prev_completed_qty or 0 for i in items))
        ws.cell(row=total_row, column=8, value=sum(i.this_completed_qty or 0 for i in items))
        ws.cell(row=total_row, column=9, value=sum(i.total_completed_qty or 0 for i in items))
        ws.cell(row=total_row, column=10, value=sum(i.this_amount or 0 for i in items))
        ws.cell(row=total_row, column=11, value=sum(i.cumulative_amount or 0 for i in items))
        for c in range(1, 15):
            ws.cell(row=total_row, column=c).border = thin_border

        col_widths = [6, 14, 40, 8, 12, 12, 12, 12, 12, 14, 14, 12, 10, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        os.makedirs(REPORT_DIR, exist_ok=True)
        filepath = os.path.join(REPORT_DIR, f"invoice_{invoice_id}.xlsx")
        wb.save(filepath)

        return send_file(
            filepath,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"PCCES_{proj.code}_計價報表.xlsx",
        )
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 分包合約管理 API（Contract）
# ═══════════════════════════════════════════════

def _recalc_contract_item(item: ContractItem):
    """計算 ContractItem 的金額 = qty × unit_price"""
    item.amount = round((item.contract_qty or 0) * (item.unit_price or 0), 2)


def _recalc_contract_amount(db: Session, contract_id: int):
    """重新計算合約金額 = 所有 items 加總"""
    total = db.query(func.coalesce(func.sum(ContractItem.amount), 0)).filter(
        ContractItem.contract_id == contract_id
    ).scalar() or 0.0
    c = db.query(Contract).filter(Contract.id == contract_id).first()
    if c:
        c.contract_amount = round(float(total), 2)
    db.flush()


def _check_contract_access(db: Session, contract_id: int, user_id: int):
    """檢查合約存取權限，回傳 (contract, project, error_response)"""
    c = db.query(Contract).filter(Contract.id == contract_id).first()
    if not c:
        return None, None, (jsonify({"detail": "合約不存在"}), 404)
    proj, err = _check_project_access(db, c.project_id, user_id)
    if err:
        return None, None, err
    return c, proj, None


# ─── Contract CRUD ───

@app.route("/api/projects/<int:project_id>/contracts/", methods=["GET"])
@require_auth
def list_contracts(project_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        contracts = db.query(Contract).filter(
            Contract.project_id == project_id
        ).order_by(Contract.created_at.desc()).all()
        return jsonify([model_to_dict(c) for c in contracts])
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/contracts/", methods=["POST"])
@require_auth
def create_contract(project_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        # 自動產生合約編號
        count = db.query(func.count(Contract.id)).filter(
            Contract.project_id == project_id
        ).scalar() or 0
        contract_no = data.get("contract_no") or f"SC-{proj.code}-{count + 1:03d}"
        c = Contract(
            project_id=project_id,
            contract_no=contract_no,
            c_name=data.get("c_name", ""),
            contractor=data.get("contractor"),
            contract_amount=data.get("contract_amount", 0),
            total_paid_amount=data.get("total_paid_amount", 0),
            total_issue_amount=data.get("total_issue_amount", 0),
            settlement_amount=data.get("settlement_amount", 0),
            status=data.get("status", "draft"),
            start_date=data.get("start_date"),
            end_date=data.get("end_date"),
            remark=data.get("remark"),
        )
        db.add(c)
        db.commit()
        db.refresh(c)
        return jsonify(model_to_dict(c)), 201
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/contracts/<int:contract_id>", methods=["GET"])
@require_auth
def get_contract(project_id, contract_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        c = db.query(Contract).filter(
            Contract.id == contract_id,
            Contract.project_id == project_id
        ).first()
        if not c:
            return jsonify({"detail": "合約不存在"}), 404
        d = model_to_dict(c)
        d["item_count"] = db.query(func.count(ContractItem.id)).filter(
            ContractItem.contract_id == contract_id
        ).scalar() or 0
        return jsonify(d)
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/contracts/<int:contract_id>", methods=["PUT"])
@require_auth
def update_contract(project_id, contract_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        c = db.query(Contract).filter(
            Contract.id == contract_id,
            Contract.project_id == project_id
        ).first()
        if not c:
            return jsonify({"detail": "合約不存在"}), 404
        if c.status not in ("draft", "active"):
            return jsonify({"detail": "僅草稿或進行中狀態可編輯"}), 400
        for key in ("contract_no", "c_name", "contractor", "contract_amount",
                     "total_paid_amount", "total_issue_amount", "settlement_amount",
                     "status", "start_date", "end_date", "remark"):
            if key in data and data[key] is not None:
                setattr(c, key, data[key])
        db.commit()
        db.refresh(c)
        return jsonify(model_to_dict(c))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/contracts/<int:contract_id>", methods=["DELETE"])
@require_auth
def delete_contract(project_id, contract_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        c = db.query(Contract).filter(
            Contract.id == contract_id,
            Contract.project_id == project_id
        ).first()
        if not c:
            return jsonify({"detail": "合約不存在"}), 404
        if c.status not in ("draft",):
            return jsonify({"detail": "僅草稿狀態可刪除"}), 400
        db.delete(c)
        db.commit()
        return jsonify({"message": "合約已刪除"})
    finally:
        db.close()


# ─── Contract Status Operations ───

@app.route("/api/projects/<int:project_id>/contracts/<int:contract_id>/close", methods=["POST"])
@require_auth
def close_contract(project_id, contract_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        c = db.query(Contract).filter(
            Contract.id == contract_id,
            Contract.project_id == project_id
        ).first()
        if not c:
            return jsonify({"detail": "合約不存在"}), 404
        if c.status not in ("active", "draft"):
            return jsonify({"detail": "僅進行中或草稿狀態可結案"}), 400
        c.status = "closed"
        db.commit()
        db.refresh(c)
        return jsonify(model_to_dict(c))
    finally:
        db.close()


@app.route("/api/projects/<int:project_id>/contracts/<int:contract_id>/finalize", methods=["POST"])
@require_auth
def finalize_contract(project_id, contract_id, user_id):
    db = next(get_db())
    try:
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err
        c = db.query(Contract).filter(
            Contract.id == contract_id,
            Contract.project_id == project_id
        ).first()
        if not c:
            return jsonify({"detail": "合約不存在"}), 404
        if c.status != "closed":
            return jsonify({"detail": "僅已結案狀態可終驗完成"}), 400
        c.status = "finalized"
        db.commit()
        db.refresh(c)
        return jsonify(model_to_dict(c))
    finally:
        db.close()


# ─── Contract Items CRUD ───

@app.route("/api/contracts/<int:contract_id>/items/", methods=["GET"])
@require_auth
def list_contract_items(contract_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        items = db.query(ContractItem).filter(
            ContractItem.contract_id == contract_id
        ).order_by(ContractItem.sort_order, ContractItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/items/", methods=["POST"])
@require_auth
def create_contract_item(contract_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        if c.status not in ("draft", "active"):
            return jsonify({"detail": "僅草稿或進行中狀態可編輯"}), 400

        item = ContractItem(
            contract_id=contract_id,
            budget_item_id=data.get("budget_item_id"),
            item_no=data.get("item_no"),
            print_no=data.get("print_no"),
            c_name=data.get("c_name"),
            c_unit=data.get("c_unit"),
            contract_qty=float(data.get("contract_qty", 0)),
            unit_price=float(data.get("unit_price", 0)),
            completed_qty=float(data.get("completed_qty", 0)),
            completed_amount=float(data.get("completed_amount", 0)),
            remark=data.get("remark"),
            sort_order=data.get("sort_order", 0),
        )
        _recalc_contract_item(item)
        db.add(item)
        db.flush()
        _recalc_contract_amount(db, contract_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_contract_item(contract_id, item_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        if c.status not in ("draft", "active"):
            return jsonify({"detail": "僅草稿或進行中狀態可編輯"}), 400
        item = db.query(ContractItem).filter(
            ContractItem.id == item_id,
            ContractItem.contract_id == contract_id
        ).first()
        if not item:
            return jsonify({"detail": "工項不存在"}), 404

        for key in ("budget_item_id", "item_no", "print_no", "c_name", "c_unit",
                     "remark", "sort_order"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])
        if "contract_qty" in data:
            item.contract_qty = float(data["contract_qty"])
        if "unit_price" in data:
            item.unit_price = float(data["unit_price"])
        if "completed_qty" in data:
            item.completed_qty = float(data["completed_qty"])
        if "completed_amount" in data:
            item.completed_amount = float(data["completed_amount"])

        _recalc_contract_item(item)
        db.flush()
        _recalc_contract_amount(db, contract_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/items/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_contract_item(contract_id, item_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        if c.status not in ("draft", "active"):
            return jsonify({"detail": "僅草稿或進行中狀態可編輯"}), 400
        item = db.query(ContractItem).filter(
            ContractItem.id == item_id,
            ContractItem.contract_id == contract_id
        ).first()
        if not item:
            return jsonify({"detail": "工項不存在"}), 404
        db.delete(item)
        db.flush()
        _recalc_contract_amount(db, contract_id)
        db.commit()
        return jsonify({"message": "工項已刪除"})
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/items/batch", methods=["POST"])
@require_auth
def batch_create_contract_items(contract_id, user_id):
    """批次從預算工項匯入合約工項"""
    data = request.get_json() or {}
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        if c.status not in ("draft", "active"):
            return jsonify({"detail": "僅草稿或進行中狀態可編輯"}), 400

        # 若傳入 budget_item_ids 則只匯入指定項目；否則匯入所有 W 類型項目
        item_ids = data.get("budget_item_ids")
        budget_query = db.query(BudgetItem).filter(
            BudgetItem.project_id == c.project_id,
            BudgetItem.kind == BudgetItemKind.W.value,
        )
        if item_ids:
            budget_query = budget_query.filter(BudgetItem.id.in_(item_ids))
        budget_items = budget_query.all()

        # 避免重複匯入
        existing_ids = set()
        existing = db.query(ContractItem.budget_item_id).filter(
            ContractItem.contract_id == contract_id,
            ContractItem.budget_item_id.isnot(None),
        ).all()
        for (eid,) in existing:
            existing_ids.add(eid)

        created = []
        for bi in budget_items:
            if bi.id in existing_ids:
                continue
            item = ContractItem(
                contract_id=contract_id,
                budget_item_id=bi.id,
                item_no=bi.item_no,
                print_no=bi.print_no,
                c_name=bi.c_name,
                c_unit=bi.c_unit,
                contract_qty=bi.quantity or 0,
                unit_price=bi.unit_price or 0,
                sort_order=0,
            )
            _recalc_contract_item(item)
            db.add(item)
            db.flush()
            created.append(model_to_dict(item))

        _recalc_contract_amount(db, contract_id)
        db.commit()
        return jsonify({"created": created, "count": len(created)}), 201
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 期別計價 API（ContractIssue）
# ═══════════════════════════════════════════════

def _calc_issue_item_amounts(item: ContractIssueItem):
    """計算 ContractIssueItem 的金額與完成率"""
    item.total_completed_qty = (item.prev_completed_qty or 0) + (item.this_completed_qty or 0)
    item.this_amount = round((item.this_completed_qty or 0) * (item.unit_price or 0), 2)
    item.cumulative_amount = round((item.total_completed_qty or 0) * (item.unit_price or 0), 2)
    item.remain_qty = max(0, (item.contract_qty or 0) - item.total_completed_qty)
    if item.contract_qty and item.contract_qty > 0:
        item.progress_rate = round((item.total_completed_qty / item.contract_qty) * 100, 2)
    else:
        item.progress_rate = 0


def _recalc_issue(db: Session, issue_id: int):
    """重新計算期別計價所有明細金額並更新主檔"""
    issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
    if not issue:
        return
    items = db.query(ContractIssueItem).filter(
        ContractIssueItem.issue_id == issue_id
    ).all()
    total_this = 0.0
    total_cum = 0.0
    for item in items:
        _calc_issue_item_amounts(item)
        total_this += item.this_amount or 0
        total_cum += item.cumulative_amount or 0
        db.flush()
    issue.total_amount = round(total_this, 2)
    issue.cumulative_amount = round(total_cum, 2)
    # 進度 = 累計金額 / 該合約總金額
    ci = db.query(ContractItem).filter(
        ContractItem.contract_id == issue.contract_id
    ).all()
    total_contract_value = sum((c.contract_qty or 0) * (c.unit_price or 0) for c in ci)
    if total_contract_value > 0:
        issue.progress_rate = round((issue.cumulative_amount / total_contract_value) * 100, 2)
    else:
        issue.progress_rate = 0
    db.flush()


# ─── Issues CRUD ───

@app.route("/api/contracts/<int:contract_id>/issues/", methods=["GET"])
@require_auth
def list_contract_issues(contract_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        issues = db.query(ContractIssue).filter(
            ContractIssue.contract_id == contract_id
        ).order_by(ContractIssue.issue_no.desc()).all()
        return jsonify([model_to_dict(i) for i in issues])
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/issues/", methods=["POST"])
@require_auth
def create_contract_issue(contract_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        # 自動決定期別編號
        last_no = db.query(func.max(ContractIssue.issue_no)).filter(
            ContractIssue.contract_id == contract_id
        ).scalar() or 0
        issue_no = last_no + 1

        issue = ContractIssue(
            contract_id=contract_id,
            issue_no=issue_no,
            c_name=data.get("c_name", f"第{issue_no}期計價"),
            status="draft",
            total_amount=0,
            remark=data.get("remark"),
            issue_date=data.get("issue_date"),
            created_by=user_id,
        )
        db.add(issue)
        db.commit()
        db.refresh(issue)
        return jsonify(model_to_dict(issue)), 201
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/issues/<int:issue_id>", methods=["GET"])
@require_auth
def get_contract_issue(contract_id, issue_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        issue = db.query(ContractIssue).filter(
            ContractIssue.id == issue_id,
            ContractIssue.contract_id == contract_id
        ).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        d = model_to_dict(issue)
        d["item_count"] = db.query(func.count(ContractIssueItem.id)).filter(
            ContractIssueItem.issue_id == issue_id
        ).scalar() or 0
        return jsonify(d)
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/issues/<int:issue_id>", methods=["PUT"])
@require_auth
def update_contract_issue(contract_id, issue_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        issue = db.query(ContractIssue).filter(
            ContractIssue.id == issue_id,
            ContractIssue.contract_id == contract_id
        ).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        for key in ("c_name", "remark", "issue_date"):
            if key in data and data[key] is not None:
                setattr(issue, key, data[key])
        db.commit()
        db.refresh(issue)
        return jsonify(model_to_dict(issue))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/issues/<int:issue_id>", methods=["DELETE"])
@require_auth
def delete_contract_issue(contract_id, issue_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        issue = db.query(ContractIssue).filter(
            ContractIssue.id == issue_id,
            ContractIssue.contract_id == contract_id
        ).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可刪除"}), 400
        db.delete(issue)
        db.commit()
        return jsonify({"message": "期別計價單已刪除"})
    finally:
        db.close()


# ─── Issue Submit / Approve ───

@app.route("/api/contracts/<int:contract_id>/issues/<int:issue_id>/submit", methods=["POST"])
@require_auth
def submit_contract_issue(contract_id, issue_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        issue = db.query(ContractIssue).filter(
            ContractIssue.id == issue_id,
            ContractIssue.contract_id == contract_id
        ).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可提交"}), 400
        # 確保至少有一筆明細
        item_count = db.query(func.count(ContractIssueItem.id)).filter(
            ContractIssueItem.issue_id == issue_id
        ).scalar() or 0
        if item_count == 0:
            return jsonify({"detail": "請先加入計價明細再提交"}), 400
        _recalc_issue(db, issue_id)
        issue.status = "submitted"
        db.commit()
        db.refresh(issue)
        return jsonify(model_to_dict(issue))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/issues/<int:issue_id>/approve", methods=["POST"])
@require_auth
def approve_contract_issue(contract_id, issue_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        issue = db.query(ContractIssue).filter(
            ContractIssue.id == issue_id,
            ContractIssue.contract_id == contract_id
        ).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        if issue.status != "submitted":
            return jsonify({"detail": "僅已提交狀態可核准"}), 400
        issue.status = "approved"
        # 核准時更新合約累計期別金額
        total_issue = db.query(func.coalesce(func.sum(ContractIssue.total_amount), 0)).filter(
            ContractIssue.contract_id == contract_id,
            ContractIssue.status == "approved"
        ).scalar() or 0.0
        c = db.query(Contract).filter(Contract.id == contract_id).first()
        if c:
            c.total_issue_amount = round(float(total_issue), 2)

        # 核准時同步更新 ContractItem 的完成數量
        issue_items = db.query(ContractIssueItem).filter(
            ContractIssueItem.issue_id == issue_id
        ).all()
        total_completed_amount = 0.0
        for ii in issue_items:
            if ii.contract_item_id:
                ci = db.query(ContractItem).filter(ContractItem.id == ii.contract_item_id).first()
                if ci:
                    ci.completed_qty = ii.total_completed_qty
                    ci.completed_amount = round(ii.total_completed_qty * ci.unit_price, 2)
                    total_completed_amount += ci.completed_amount
        if c:
            c.total_paid_amount = round(total_completed_amount, 2)

        db.commit()
        db.refresh(issue)
        return jsonify(model_to_dict(issue))
    finally:
        db.close()


# ─── Issue Items CRUD ───

@app.route("/api/issues/<int:issue_id>/items/", methods=["GET"])
@require_auth
def list_issue_items(issue_id, user_id):
    db = next(get_db())
    try:
        issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        c, proj, err = _check_contract_access(db, issue.contract_id, user_id)
        if err:
            return err
        items = db.query(ContractIssueItem).filter(
            ContractIssueItem.issue_id == issue_id
        ).order_by(ContractIssueItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/issues/<int:issue_id>/items/", methods=["POST"])
@require_auth
def create_issue_item(issue_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        c, proj, err = _check_contract_access(db, issue.contract_id, user_id)
        if err:
            return err
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        contract_qty = float(data.get("contract_qty", 0))
        unit_price = float(data.get("unit_price", 0))
        prev_completed = float(data.get("prev_completed_qty", 0))
        this_completed = float(data.get("this_completed_qty", 0))

        item = ContractIssueItem(
            issue_id=issue_id,
            contract_item_id=data.get("contract_item_id"),
            c_name=data.get("c_name"),
            c_unit=data.get("c_unit"),
            contract_qty=contract_qty,
            unit_price=unit_price,
            prev_completed_qty=prev_completed,
            this_completed_qty=this_completed,
            remark=data.get("remark"),
        )
        # 超量完成檢查
        total_check = prev_completed + this_completed
        if total_check > contract_qty:
            return jsonify({"detail": f"累計完成數量 ({total_check:.2f}) 超過合約數量 ({contract_qty:.2f})"}), 400
        _calc_issue_item_amounts(item)
        db.add(item)
        db.flush()
        _recalc_issue(db, issue_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/issues/<int:issue_id>/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_issue_item(issue_id, item_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        c, proj, err = _check_contract_access(db, issue.contract_id, user_id)
        if err:
            return err
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        item = db.query(ContractIssueItem).filter(
            ContractIssueItem.id == item_id,
            ContractIssueItem.issue_id == issue_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404

        for key in ("this_completed_qty", "prev_completed_qty", "contract_qty",
                     "unit_price", "remark", "c_name", "c_unit"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])

        # 超量完成檢查
        new_total = (item.prev_completed_qty or 0) + (item.this_completed_qty or 0)
        if new_total > (item.contract_qty or 0):
            return jsonify({"detail": f"累計完成數量 ({new_total:.2f}) 超過合約數量 ({item.contract_qty:.2f})"}), 400

        _calc_issue_item_amounts(item)
        db.flush()
        _recalc_issue(db, issue_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/issues/<int:issue_id>/items/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_issue_item(issue_id, item_id, user_id):
    db = next(get_db())
    try:
        issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        c, proj, err = _check_contract_access(db, issue.contract_id, user_id)
        if err:
            return err
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        item = db.query(ContractIssueItem).filter(
            ContractIssueItem.id == item_id,
            ContractIssueItem.issue_id == issue_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404
        db.delete(item)
        db.flush()
        _recalc_issue(db, issue_id)
        db.commit()
        return jsonify({"message": "明細已刪除"})
    finally:
        db.close()


@app.route("/api/issues/<int:issue_id>/items/recalc", methods=["POST"])
@require_auth
def recalc_issue_items(issue_id, user_id):
    db = next(get_db())
    try:
        issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        c, proj, err = _check_contract_access(db, issue.contract_id, user_id)
        if err:
            return err
        _recalc_issue(db, issue_id)
        db.commit()
        return jsonify({"message": "期別金額重新計算完成"})
    finally:
        db.close()


@app.route("/api/issues/<int:issue_id>/items/batch-from-contract", methods=["POST"])
@require_auth
def batch_issue_items_from_contract(issue_id, user_id):
    """批次從合約工項導入期別計價明細"""
    db = next(get_db())
    try:
        issue = db.query(ContractIssue).filter(ContractIssue.id == issue_id).first()
        if not issue:
            return jsonify({"detail": "期別計價單不存在"}), 404
        c, proj, err = _check_contract_access(db, issue.contract_id, user_id)
        if err:
            return err
        if issue.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        # 取得前一期以獲取 prev_completed_qty
        prev_issue = db.query(ContractIssue).filter(
            ContractIssue.contract_id == issue.contract_id,
            ContractIssue.issue_no < issue.issue_no,
            ContractIssue.status == "approved"
        ).order_by(ContractIssue.issue_no.desc()).first()

        # 建立前一期完成數量映射
        prev_completed_map = {}
        if prev_issue:
            prev_items = db.query(ContractIssueItem).filter(
                ContractIssueItem.issue_id == prev_issue.id
            ).all()
            for pi in prev_items:
                if pi.contract_item_id:
                    prev_completed_map[pi.contract_item_id] = pi.total_completed_qty or 0

        contract_items = db.query(ContractItem).filter(
            ContractItem.contract_id == issue.contract_id
        ).order_by(ContractItem.sort_order, ContractItem.id).all()

        # 避免重複導入
        existing_ids = set()
        existing = db.query(ContractIssueItem.contract_item_id).filter(
            ContractIssueItem.issue_id == issue_id,
            ContractIssueItem.contract_item_id.isnot(None),
        ).all()
        for (eid,) in existing:
            existing_ids.add(eid)

        created = []
        for ci in contract_items:
            if ci.id in existing_ids:
                continue
            prev_qty = prev_completed_map.get(ci.id, 0)
            item = ContractIssueItem(
                issue_id=issue_id,
                contract_item_id=ci.id,
                c_name=ci.c_name,
                c_unit=ci.c_unit,
                contract_qty=ci.contract_qty or 0,
                unit_price=ci.unit_price or 0,
                prev_completed_qty=prev_qty,
                this_completed_qty=0,
            )
            _calc_issue_item_amounts(item)
            db.add(item)
            db.flush()
            created.append(model_to_dict(item))

        _recalc_issue(db, issue_id)
        db.commit()
        return jsonify({"created": created, "count": len(created)}), 201
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 結算管理 API（ContractSettlement）
# ═══════════════════════════════════════════════

def _recalc_settlement(db: Session, settlement_id: int):
    """重新計算結算金額"""
    st = db.query(ContractSettlement).filter(ContractSettlement.id == settlement_id).first()
    if not st:
        return
    items = db.query(ContractSettlementItem).filter(
        ContractSettlementItem.settlement_id == settlement_id
    ).all()
    total_add = 0.0
    total_deduct = 0.0
    for item in items:
        # actual_amount = actual_qty × actual_unit_price
        item.actual_amount = round((item.actual_qty or 0) * (item.actual_unit_price or 0), 2)
        # diff_amount = actual_amount - contract_amount
        item.diff_amount = round((item.actual_amount or 0) - (item.contract_amount or 0), 2)
        if item.diff_amount > 0:
            total_add += item.diff_amount
        else:
            total_deduct += abs(item.diff_amount)
        db.flush()
    st.total_add_amount = round(total_add, 2)
    st.total_deduct_amount = round(total_deduct, 2)
    # 從明細 actual_amount 加總（避免 header snapshot 不一致）
    total_actual = db.query(func.coalesce(func.sum(ContractSettlementItem.actual_amount), 0)).filter(
        ContractSettlementItem.settlement_id == settlement_id
    ).scalar() or 0.0
    st.settlement_amount = round(float(total_actual), 2)
    db.flush()


# ─── Settlements CRUD ───

@app.route("/api/contracts/<int:contract_id>/settlements/", methods=["GET"])
@require_auth
def list_contract_settlements(contract_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        items = db.query(ContractSettlement).filter(
            ContractSettlement.contract_id == contract_id
        ).order_by(ContractSettlement.id.desc()).all()
        return jsonify([model_to_dict(s) for s in items])
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/settlements/", methods=["POST"])
@require_auth
def create_contract_settlement(contract_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        count = db.query(func.count(ContractSettlement.id)).filter(
            ContractSettlement.contract_id == contract_id
        ).scalar() or 0
        settlement_no = data.get("settlement_no") or f"ST-{count + 1:03d}"

        st = ContractSettlement(
            contract_id=contract_id,
            settlement_no=settlement_no,
            c_name=data.get("c_name", "結算"),
            settlement_date=data.get("settlement_date"),
            contract_amount=c.contract_amount or 0,
            remark=data.get("remark"),
            status="draft",
            created_by=user_id,
        )
        db.add(st)
        db.commit()
        db.refresh(st)
        return jsonify(model_to_dict(st)), 201
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/settlements/<int:settlement_id>", methods=["GET"])
@require_auth
def get_contract_settlement(contract_id, settlement_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        st = db.query(ContractSettlement).filter(
            ContractSettlement.id == settlement_id,
            ContractSettlement.contract_id == contract_id
        ).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        return jsonify(model_to_dict(st))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/settlements/<int:settlement_id>", methods=["PUT"])
@require_auth
def update_contract_settlement(contract_id, settlement_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        st = db.query(ContractSettlement).filter(
            ContractSettlement.id == settlement_id,
            ContractSettlement.contract_id == contract_id
        ).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        if st.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        for key in ("settlement_no", "c_name", "settlement_date", "remark"):
            if key in data and data[key] is not None:
                setattr(st, key, data[key])
        db.commit()
        db.refresh(st)
        return jsonify(model_to_dict(st))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/settlements/<int:settlement_id>", methods=["DELETE"])
@require_auth
def delete_contract_settlement(contract_id, settlement_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        st = db.query(ContractSettlement).filter(
            ContractSettlement.id == settlement_id,
            ContractSettlement.contract_id == contract_id
        ).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        if st.status != "draft":
            return jsonify({"detail": "僅草稿狀態可刪除"}), 400
        db.delete(st)
        db.commit()
        return jsonify({"message": "結算單已刪除"})
    finally:
        db.close()


# ─── Settlement Submit / Approve ───

@app.route("/api/contracts/<int:contract_id>/settlements/<int:settlement_id>/submit", methods=["POST"])
@require_auth
def submit_contract_settlement(contract_id, settlement_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        st = db.query(ContractSettlement).filter(
            ContractSettlement.id == settlement_id,
            ContractSettlement.contract_id == contract_id
        ).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        if st.status != "draft":
            return jsonify({"detail": "僅草稿狀態可提交"}), 400
        _recalc_settlement(db, settlement_id)
        st.status = "submitted"
        db.commit()
        db.refresh(st)
        return jsonify(model_to_dict(st))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/settlements/<int:settlement_id>/approve", methods=["POST"])
@require_auth
def approve_contract_settlement(contract_id, settlement_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        st = db.query(ContractSettlement).filter(
            ContractSettlement.id == settlement_id,
            ContractSettlement.contract_id == contract_id
        ).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        if st.status != "submitted":
            return jsonify({"detail": "僅已提交狀態可核准"}), 400
        st.status = "approved"
        # 核准時更新合約結算金額
        c = db.query(Contract).filter(Contract.id == contract_id).first()
        if c:
            c.settlement_amount = st.settlement_amount or 0
        db.commit()
        db.refresh(st)
        return jsonify(model_to_dict(st))
    finally:
        db.close()


# ─── Settlement Items CRUD ───

@app.route("/api/settlements/<int:settlement_id>/items/", methods=["GET"])
@require_auth
def list_settlement_items(settlement_id, user_id):
    db = next(get_db())
    try:
        st = db.query(ContractSettlement).filter(ContractSettlement.id == settlement_id).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        c, proj, err = _check_contract_access(db, st.contract_id, user_id)
        if err:
            return err
        items = db.query(ContractSettlementItem).filter(
            ContractSettlementItem.settlement_id == settlement_id
        ).order_by(ContractSettlementItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/settlements/<int:settlement_id>/items/", methods=["POST"])
@require_auth
def create_settlement_item(settlement_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        st = db.query(ContractSettlement).filter(ContractSettlement.id == settlement_id).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        c, proj, err = _check_contract_access(db, st.contract_id, user_id)
        if err:
            return err
        if st.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        contract_qty = float(data.get("contract_qty", 0))
        contract_up = float(data.get("contract_unit_price", 0))
        actual_qty = float(data.get("actual_qty", 0))
        actual_up = float(data.get("actual_unit_price", 0))

        item = ContractSettlementItem(
            settlement_id=settlement_id,
            budget_item_id=data.get("budget_item_id"),
            c_name=data.get("c_name"),
            c_unit=data.get("c_unit"),
            contract_qty=contract_qty,
            contract_unit_price=contract_up,
            contract_amount=round(contract_qty * contract_up, 2),
            actual_qty=actual_qty,
            actual_unit_price=actual_up,
            actual_amount=round(actual_qty * actual_up, 2),
            diff_amount=0,
            remark=data.get("remark"),
        )
        item.diff_amount = round((item.actual_amount or 0) - (item.contract_amount or 0), 2)
        db.add(item)
        db.flush()
        _recalc_settlement(db, settlement_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/settlements/<int:settlement_id>/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_settlement_item(settlement_id, item_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        st = db.query(ContractSettlement).filter(ContractSettlement.id == settlement_id).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        c, proj, err = _check_contract_access(db, st.contract_id, user_id)
        if err:
            return err
        if st.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        item = db.query(ContractSettlementItem).filter(
            ContractSettlementItem.id == item_id,
            ContractSettlementItem.settlement_id == settlement_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404

        for key in ("contract_qty", "contract_unit_price", "actual_qty",
                     "actual_unit_price", "c_name", "c_unit", "remark"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])

        # 重新計算金額
        item.contract_amount = round((item.contract_qty or 0) * (item.contract_unit_price or 0), 2)
        item.actual_amount = round((item.actual_qty or 0) * (item.actual_unit_price or 0), 2)
        item.diff_amount = round((item.actual_amount or 0) - (item.contract_amount or 0), 2)

        db.flush()
        _recalc_settlement(db, settlement_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/settlements/<int:settlement_id>/items/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_settlement_item(settlement_id, item_id, user_id):
    db = next(get_db())
    try:
        st = db.query(ContractSettlement).filter(ContractSettlement.id == settlement_id).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        c, proj, err = _check_contract_access(db, st.contract_id, user_id)
        if err:
            return err
        if st.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        item = db.query(ContractSettlementItem).filter(
            ContractSettlementItem.id == item_id,
            ContractSettlementItem.settlement_id == settlement_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404
        db.delete(item)
        db.flush()
        _recalc_settlement(db, settlement_id)
        db.commit()
        return jsonify({"message": "明細已刪除"})
    finally:
        db.close()


@app.route("/api/settlements/<int:settlement_id>/items/recalc", methods=["POST"])
@require_auth
def recalc_settlement_items(settlement_id, user_id):
    db = next(get_db())
    try:
        st = db.query(ContractSettlement).filter(ContractSettlement.id == settlement_id).first()
        if not st:
            return jsonify({"detail": "結算單不存在"}), 404
        c, proj, err = _check_contract_access(db, st.contract_id, user_id)
        if err:
            return err
        _recalc_settlement(db, settlement_id)
        db.commit()
        return jsonify({"message": "結算金額重新計算完成"})
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 終驗管理 API（ContractFinalAcceptance）
# ═══════════════════════════════════════════════

def _recalc_acceptance(db: Session, acceptance_id: int):
    """重新計算終驗統計"""
    ac = db.query(ContractFinalAcceptance).filter(
        ContractFinalAcceptance.id == acceptance_id
    ).first()
    if not ac:
        return
    items = db.query(ContractFinalAcceptanceItem).filter(
        ContractFinalAcceptanceItem.acceptance_id == acceptance_id
    ).all()
    total_contract = sum(i.contract_qty or 0 for i in items)
    total_accepted = sum(i.accepted_qty or 0 for i in items)
    total_rejected = sum(i.rejected_qty or 0 for i in items)
    # 結果回寫到備註統計中（前端可讀取）
    db.flush()


# ─── Acceptances CRUD ───

@app.route("/api/contracts/<int:contract_id>/acceptances/", methods=["GET"])
@require_auth
def list_contract_acceptances(contract_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        items = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.contract_id == contract_id
        ).order_by(ContractFinalAcceptance.id.desc()).all()
        return jsonify([model_to_dict(a) for a in items])
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/acceptances/", methods=["POST"])
@require_auth
def create_contract_acceptance(contract_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        count = db.query(func.count(ContractFinalAcceptance.id)).filter(
            ContractFinalAcceptance.contract_id == contract_id
        ).scalar() or 0
        acceptance_no = data.get("acceptance_no") or f"FA-{count + 1:03d}"

        ac = ContractFinalAcceptance(
            contract_id=contract_id,
            acceptance_no=acceptance_no,
            c_name=data.get("c_name", "終驗"),
            acceptance_date=data.get("acceptance_date"),
            inspector=data.get("inspector"),
            result=data.get("result"),
            defect_description=data.get("defect_description"),
            remark=data.get("remark"),
            status="draft",
            created_by=user_id,
        )
        db.add(ac)
        db.commit()
        db.refresh(ac)
        return jsonify(model_to_dict(ac)), 201
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/acceptances/<int:acceptance_id>", methods=["GET"])
@require_auth
def get_contract_acceptance(contract_id, acceptance_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id,
            ContractFinalAcceptance.contract_id == contract_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        return jsonify(model_to_dict(ac))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/acceptances/<int:acceptance_id>", methods=["PUT"])
@require_auth
def update_contract_acceptance(contract_id, acceptance_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id,
            ContractFinalAcceptance.contract_id == contract_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        for key in ("acceptance_no", "c_name", "acceptance_date", "inspector",
                     "result", "defect_description", "remark"):
            if key in data and data[key] is not None:
                setattr(ac, key, data[key])
        db.commit()
        db.refresh(ac)
        return jsonify(model_to_dict(ac))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/acceptances/<int:acceptance_id>", methods=["DELETE"])
@require_auth
def delete_contract_acceptance(contract_id, acceptance_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id,
            ContractFinalAcceptance.contract_id == contract_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可刪除"}), 400
        db.delete(ac)
        db.commit()
        return jsonify({"message": "終驗單已刪除"})
    finally:
        db.close()


# ─── Acceptance Submit / Approve ───

@app.route("/api/contracts/<int:contract_id>/acceptances/<int:acceptance_id>/submit", methods=["POST"])
@require_auth
def submit_contract_acceptance(contract_id, acceptance_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id,
            ContractFinalAcceptance.contract_id == contract_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可提交"}), 400
        ac.status = "submitted"
        db.commit()
        db.refresh(ac)
        return jsonify(model_to_dict(ac))
    finally:
        db.close()


@app.route("/api/contracts/<int:contract_id>/acceptances/<int:acceptance_id>/approve", methods=["POST"])
@require_auth
def approve_contract_acceptance(contract_id, acceptance_id, user_id):
    db = next(get_db())
    try:
        c, proj, err = _check_contract_access(db, contract_id, user_id)
        if err:
            return err
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id,
            ContractFinalAcceptance.contract_id == contract_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        if ac.status != "submitted":
            return jsonify({"detail": "僅已提交狀態可核准"}), 400
        ac.status = "approved"
        db.commit()
        db.refresh(ac)
        return jsonify(model_to_dict(ac))
    finally:
        db.close()


# ─── Acceptance Items CRUD ───

@app.route("/api/acceptances/<int:acceptance_id>/items/", methods=["GET"])
@require_auth
def list_acceptance_items(acceptance_id, user_id):
    db = next(get_db())
    try:
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        c, proj, err = _check_contract_access(db, ac.contract_id, user_id)
        if err:
            return err
        items = db.query(ContractFinalAcceptanceItem).filter(
            ContractFinalAcceptanceItem.acceptance_id == acceptance_id
        ).order_by(ContractFinalAcceptanceItem.id).all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


@app.route("/api/acceptances/<int:acceptance_id>/items/", methods=["POST"])
@require_auth
def create_acceptance_item(acceptance_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        c, proj, err = _check_contract_access(db, ac.contract_id, user_id)
        if err:
            return err
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        item = ContractFinalAcceptanceItem(
            acceptance_id=acceptance_id,
            budget_item_id=data.get("budget_item_id"),
            c_name=data.get("c_name"),
            c_unit=data.get("c_unit"),
            contract_qty=float(data.get("contract_qty", 0)),
            actual_qty=float(data.get("actual_qty", 0)),
            accepted_qty=float(data.get("accepted_qty", 0)),
            rejected_qty=float(data.get("rejected_qty", 0)),
            remark=data.get("remark"),
        )
        db.add(item)
        db.flush()
        _recalc_acceptance(db, acceptance_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/acceptances/<int:acceptance_id>/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_acceptance_item(acceptance_id, item_id, user_id):
    data = request.get_json()
    db = next(get_db())
    try:
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        c, proj, err = _check_contract_access(db, ac.contract_id, user_id)
        if err:
            return err
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        item = db.query(ContractFinalAcceptanceItem).filter(
            ContractFinalAcceptanceItem.id == item_id,
            ContractFinalAcceptanceItem.acceptance_id == acceptance_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404

        for key in ("contract_qty", "actual_qty", "accepted_qty", "rejected_qty",
                     "c_name", "c_unit", "remark"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])

        db.flush()
        _recalc_acceptance(db, acceptance_id)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/acceptances/<int:acceptance_id>/items/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_acceptance_item(acceptance_id, item_id, user_id):
    db = next(get_db())
    try:
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        c, proj, err = _check_contract_access(db, ac.contract_id, user_id)
        if err:
            return err
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400
        item = db.query(ContractFinalAcceptanceItem).filter(
            ContractFinalAcceptanceItem.id == item_id,
            ContractFinalAcceptanceItem.acceptance_id == acceptance_id
        ).first()
        if not item:
            return jsonify({"detail": "明細不存在"}), 404
        db.delete(item)
        db.flush()
        _recalc_acceptance(db, acceptance_id)
        db.commit()
        return jsonify({"message": "明細已刪除"})
    finally:
        db.close()


@app.route("/api/acceptances/<int:acceptance_id>/items/recalc", methods=["POST"])
@require_auth
def recalc_acceptance_items(acceptance_id, user_id):
    db = next(get_db())
    try:
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        c, proj, err = _check_contract_access(db, ac.contract_id, user_id)
        if err:
            return err
        _recalc_acceptance(db, acceptance_id)
        db.commit()
        return jsonify({"message": "終驗金額重新計算完成"})
    finally:
        db.close()


@app.route("/api/acceptances/<int:acceptance_id>/items/batch-from-contract", methods=["POST"])
@require_auth
def batch_acceptance_items_from_contract(acceptance_id, user_id):
    """批次從合約工項導入終驗明細"""
    db = next(get_db())
    try:
        ac = db.query(ContractFinalAcceptance).filter(
            ContractFinalAcceptance.id == acceptance_id
        ).first()
        if not ac:
            return jsonify({"detail": "終驗單不存在"}), 404
        c, proj, err = _check_contract_access(db, ac.contract_id, user_id)
        if err:
            return err
        if ac.status != "draft":
            return jsonify({"detail": "僅草稿狀態可編輯"}), 400

        contract_items = db.query(ContractItem).filter(
            ContractItem.contract_id == ac.contract_id
        ).order_by(ContractItem.sort_order, ContractItem.id).all()

        existing_ids = set()
        existing = db.query(ContractFinalAcceptanceItem.budget_item_id).filter(
            ContractFinalAcceptanceItem.acceptance_id == acceptance_id,
            ContractFinalAcceptanceItem.budget_item_id.isnot(None),
        ).all()
        for (eid,) in existing:
            existing_ids.add(eid)

        created = []
        for ci in contract_items:
            if ci.budget_item_id and ci.budget_item_id in existing_ids:
                continue
            item = ContractFinalAcceptanceItem(
                acceptance_id=acceptance_id,
                budget_item_id=ci.budget_item_id,
                c_name=ci.c_name,
                c_unit=ci.c_unit,
                contract_qty=ci.contract_qty or 0,
                actual_qty=ci.completed_qty or 0,
                accepted_qty=ci.completed_qty or 0,
                rejected_qty=0,
            )
            db.add(item)
            db.flush()
            created.append(model_to_dict(item))

        _recalc_acceptance(db, acceptance_id)
        db.commit()
        return jsonify({"created": created, "count": len(created)}), 201
    finally:
        db.close()


# ═══════════════════════════════════════════════
# MrsBase 公共單價庫 API
# ═══════════════════════════════════════════════

# ── 分類（Category）API ──

def _build_category_tree(db, parent_id=None):
    """遞迴建立分類樹（巢狀 JSON）"""
    query = db.query(MrsBaseCategory)
    if parent_id is None:
        query = query.filter(MrsBaseCategory.parent_id.is_(None))
    else:
        query = query.filter(MrsBaseCategory.parent_id == parent_id)
    query = query.order_by(MrsBaseCategory.sort_order, MrsBaseCategory.id)
    cats = query.all()
    result = []
    for cat in cats:
        d = model_to_dict(cat)
        d["children"] = _build_category_tree(db, cat.id)
        # 計算此分類下的項目數量
        d["item_count"] = db.query(func.count(MrsBaseItem.id)).filter(
            MrsBaseItem.category_id == cat.id
        ).scalar() or 0
        result.append(d)
    return result


@app.route("/api/mrs-base/categories", methods=["GET"])
@require_auth
def get_mrs_base_categories(user_id):
    """取得分類樹（巢狀 JSON）"""
    db = next(get_db())
    try:
        tree = _build_category_tree(db)
        return jsonify(tree)
    finally:
        db.close()


@app.route("/api/mrs-base/categories/flat", methods=["GET"])
@require_auth
def get_mrs_base_categories_flat(user_id):
    """取得分類平面列表"""
    db = next(get_db())
    try:
        cats = db.query(MrsBaseCategory).order_by(MrsBaseCategory.sort_order, MrsBaseCategory.id).all()
        return jsonify([model_to_dict(c) for c in cats])
    finally:
        db.close()


@app.route("/api/mrs-base/categories", methods=["POST"])
@require_auth
def create_mrs_base_category(user_id):
    """建立分類"""
    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供資料"}), 400
    db = next(get_db())
    try:
        parent_id = data.get("parent_id")
        level_no = 0
        if parent_id:
            parent = db.query(MrsBaseCategory).filter(MrsBaseCategory.id == parent_id).first()
            if not parent:
                return jsonify({"detail": "父分類不存在"}), 404
            level_no = (parent.level_no or 0) + 1

        cat = MrsBaseCategory(
            parent_id=parent_id,
            code=data.get("code", ""),
            c_name=data["c_name"],
            sort_order=data.get("sort_order", 0),
            level_no=level_no,
        )
        db.add(cat)
        db.commit()
        db.refresh(cat)
        return jsonify(model_to_dict(cat)), 201
    finally:
        db.close()


@app.route("/api/mrs-base/categories/<int:cat_id>", methods=["PUT"])
@require_auth
def update_mrs_base_category(cat_id, user_id):
    """更新分類"""
    data = request.get_json()
    db = next(get_db())
    try:
        cat = db.query(MrsBaseCategory).filter(MrsBaseCategory.id == cat_id).first()
        if not cat:
            return jsonify({"detail": "分類不存在"}), 404
        for key in ("code", "c_name", "sort_order"):
            if key in data and data[key] is not None:
                setattr(cat, key, data[key])
        if "parent_id" in data:
            parent_id = data["parent_id"]
            cat.parent_id = parent_id
            if parent_id:
                parent = db.query(MrsBaseCategory).filter(MrsBaseCategory.id == parent_id).first()
                cat.level_no = (parent.level_no or 0) + 1 if parent else 0
            else:
                cat.level_no = 0
        db.commit()
        db.refresh(cat)
        return jsonify(model_to_dict(cat))
    finally:
        db.close()


@app.route("/api/mrs-base/categories/<int:cat_id>", methods=["DELETE"])
@require_auth
def delete_mrs_base_category(cat_id, user_id):
    """刪除分類（檢查無子分類及項目）"""
    db = next(get_db())
    try:
        cat = db.query(MrsBaseCategory).filter(MrsBaseCategory.id == cat_id).first()
        if not cat:
            return jsonify({"detail": "分類不存在"}), 404
        # 檢查子分類
        child_count = db.query(func.count(MrsBaseCategory.id)).filter(
            MrsBaseCategory.parent_id == cat_id
        ).scalar() or 0
        if child_count > 0:
            return jsonify({"detail": "此分類下有子分類，無法刪除"}), 400
        # 檢查項目
        item_count = db.query(func.count(MrsBaseItem.id)).filter(
            MrsBaseItem.category_id == cat_id
        ).scalar() or 0
        if item_count > 0:
            return jsonify({"detail": "此分類下有公共單價項目，無法刪除"}), 400
        db.delete(cat)
        db.commit()
        return jsonify({"message": "分類已刪除"})
    finally:
        db.close()


# ── 項目（Item）API ──

@app.route("/api/mrs-base/items", methods=["GET"])
@require_auth
def list_mrs_base_items(user_id):
    """列表（支援查詢參數：category_id, q, kind, page, per_page）"""
    db = next(get_db())
    try:
        query = db.query(MrsBaseItem)

        # 篩選
        category_id = request.args.get("category_id")
        if category_id:
            query = query.filter(MrsBaseItem.category_id == int(category_id))

        q = request.args.get("q")
        if q:
            pattern = f"%{q}%"
            query = query.filter(
                MrsBaseItem.code.ilike(pattern) | MrsBaseItem.c_name.ilike(pattern)
            )

        kind = request.args.get("kind")
        if kind:
            query = query.filter(MrsBaseItem.cost_kind == kind)

        approved = request.args.get("approved")
        if approved is not None:
            is_approved = approved.lower() in ("1", "true", "yes")
            query = query.filter(MrsBaseItem.is_approved == is_approved)

        # 排序
        query = query.order_by(MrsBaseItem.code, MrsBaseItem.id)

        # 分頁
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 50))
        total = query.count()
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            "items": [model_to_dict(i) for i in items],
            "total": total,
            "page": page,
            "per_page": per_page,
        })
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>", methods=["GET"])
@require_auth
def get_mrs_base_item(item_id, user_id):
    """單筆（含 breakdown_items）"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        d = model_to_dict(item)
        # 帶入工料機組成
        breakdowns = db.query(MrsBaseBreakdownItem).filter(
            MrsBaseBreakdownItem.item_id == item_id
        ).order_by(MrsBaseBreakdownItem.id).all()
        d["breakdown_items"] = [model_to_dict(b) for b in breakdowns]
        d["breakdown_total"] = sum((b.amount or 0) for b in breakdowns)
        return jsonify(d)
    finally:
        db.close()


@app.route("/api/mrs-base/items", methods=["POST"])
@require_auth
def create_mrs_base_item(user_id):
    """新增項目"""
    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供資料"}), 400
    db = next(get_db())
    try:
        # 檢查編碼唯一性
        existing = db.query(MrsBaseItem).filter(MrsBaseItem.code == data["code"]).first()
        if existing:
            return jsonify({"detail": f"編碼 '{data['code']}' 已存在"}), 409

        item = MrsBaseItem(
            category_id=data["category_id"],
            code=data["code"],
            pub_code=data.get("pub_code"),
            c_name=data["c_name"],
            e_name=data.get("e_name"),
            c_unit=data.get("c_unit", "式"),
            e_unit=data.get("e_unit"),
            unit_price=data.get("unit_price", 0),
            cost_kind=data.get("cost_kind", "料"),
            item_type=data.get("item_type", "W"),
            is_analysis=data.get("is_analysis", False),
            labor_rate=data.get("labor_rate", 0),
            material_rate=data.get("material_rate", 0),
            equipment_rate=data.get("equipment_rate", 0),
            misc_rate=data.get("misc_rate", 0),
            decimal_qty=data.get("decimal_qty", 2),
            decimal_price=data.get("decimal_price", 2),
            decimal_amount=data.get("decimal_amount", 2),
            memo=data.get("memo"),
            is_approved=data.get("is_approved", False),
            created_by=user_id,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>", methods=["PUT"])
@require_auth
def update_mrs_base_item(item_id, user_id):
    """更新項目"""
    data = request.get_json()
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404

        # 檢查編碼唯一性（若變更）
        if "code" in data and data["code"] != item.code:
            existing = db.query(MrsBaseItem).filter(MrsBaseItem.code == data["code"]).first()
            if existing:
                return jsonify({"detail": f"編碼 '{data['code']}' 已存在"}), 409

        for key in ("category_id", "code", "pub_code", "c_name", "e_name",
                     "c_unit", "e_unit", "unit_price", "cost_kind", "item_type",
                     "is_analysis", "labor_rate", "material_rate", "equipment_rate",
                     "misc_rate", "decimal_qty", "decimal_price", "decimal_amount",
                     "memo", "is_approved"):
            if key in data and data[key] is not None:
                setattr(item, key, data[key])

        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>", methods=["DELETE"])
@require_auth
def delete_mrs_base_item(item_id, user_id):
    """刪除項目（含 CASCADE 子項）"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        db.delete(item)
        db.commit()
        return jsonify({"message": "項目已刪除"})
    finally:
        db.close()


@app.route("/api/mrs-base/items/batch-delete", methods=["POST"])
@require_auth
def batch_delete_mrs_base_items(user_id):
    """批次刪除項目"""
    data = request.get_json()
    if not data or "ids" not in data:
        return jsonify({"detail": "請提供 ids 陣列"}), 400
    db = next(get_db())
    try:
        ids = data["ids"]
        deleted = 0
        for item_id in ids:
            item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
            if item:
                db.delete(item)
                deleted += 1
        db.commit()
        return jsonify({"message": f"已刪除 {deleted} 筆項目", "deleted": deleted})
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>/approve", methods=["POST"])
@require_auth
def approve_mrs_base_item(item_id, user_id):
    """審核通過"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        item.is_approved = True
        item.approved_by = user_id
        item.approved_at = datetime.now(timezone.utc)
        db.commit()
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>/unapprove", methods=["POST"])
@require_auth
def unapprove_mrs_base_item(item_id, user_id):
    """取消審核"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        item.is_approved = False
        item.approved_by = None
        item.approved_at = None
        db.commit()
        return jsonify(model_to_dict(item))
    finally:
        db.close()


# ── 工料機組成（Breakdown）API ──

@app.route("/api/mrs-base/items/<int:item_id>/breakdown", methods=["GET"])
@require_auth
def list_mrs_base_breakdown(item_id, user_id):
    """取得工料機組成列表"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        items = db.query(MrsBaseBreakdownItem).filter(
            MrsBaseBreakdownItem.item_id == item_id
        ).order_by(MrsBaseBreakdownItem.id).all()
        return jsonify([model_to_dict(b) for b in items])
    finally:
        db.close()


def _recalc_mrs_base_item(db, item_id):
    """重新計算 MrsBase 項目的單價（工料機組成加總）"""
    total = db.query(func.coalesce(func.sum(MrsBaseBreakdownItem.amount), 0)).filter(
        MrsBaseBreakdownItem.item_id == item_id
    ).scalar() or 0.0
    item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
    if item:
        item.unit_price = round(float(total), 2)
    return float(total)


@app.route("/api/mrs-base/items/<int:item_id>/breakdown", methods=["POST"])
@require_auth
def create_mrs_base_breakdown(item_id, user_id):
    """新增工料機組成細項（自動更新 item.unit_price）"""
    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供資料"}), 400
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404

        qty = float(data.get("quantity", 0))
        up = float(data.get("unit_price", 0))
        bd = MrsBaseBreakdownItem(
            item_id=item_id,
            code=data.get("code", ""),
            c_name=data.get("c_name", ""),
            c_unit=data.get("c_unit", "式"),
            quantity=qty,
            unit_price=up,
            amount=round(qty * up, 2),
            category=data.get("category", "material"),
            remark=data.get("remark"),
        )
        db.add(bd)
        db.flush()

        # 自動重新計算單價
        _recalc_mrs_base_item(db, item_id)
        db.commit()
        db.refresh(bd)
        return jsonify(model_to_dict(bd)), 201
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>/breakdown/<int:bd_id>", methods=["PUT"])
@require_auth
def update_mrs_base_breakdown(item_id, bd_id, user_id):
    """更新工料機組成細項"""
    data = request.get_json()
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        bd = db.query(MrsBaseBreakdownItem).filter(
            MrsBaseBreakdownItem.id == bd_id,
            MrsBaseBreakdownItem.item_id == item_id
        ).first()
        if not bd:
            return jsonify({"detail": "細項不存在"}), 404

        for key in ("code", "c_name", "c_unit", "quantity", "unit_price", "category", "remark"):
            if key in data and data[key] is not None:
                setattr(bd, key, data[key])

        # 重新計算金額
        bd.amount = round((bd.quantity or 0) * (bd.unit_price or 0), 2)

        db.flush()
        # 自動重新計算單價
        _recalc_mrs_base_item(db, item_id)
        db.commit()
        db.refresh(bd)
        return jsonify(model_to_dict(bd))
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>/breakdown/<int:bd_id>", methods=["DELETE"])
@require_auth
def delete_mrs_base_breakdown(item_id, bd_id, user_id):
    """刪除工料機組成細項"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        bd = db.query(MrsBaseBreakdownItem).filter(
            MrsBaseBreakdownItem.id == bd_id,
            MrsBaseBreakdownItem.item_id == item_id
        ).first()
        if not bd:
            return jsonify({"detail": "細項不存在"}), 404
        db.delete(bd)
        db.flush()
        # 自動重新計算單價
        _recalc_mrs_base_item(db, item_id)
        db.commit()
        return jsonify({"message": "細項已刪除"})
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>/breakdown/recalc", methods=["POST"])
@require_auth
def recalc_mrs_base_breakdown(item_id, user_id):
    """重新計算所有工料機組成加總"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        total = _recalc_mrs_base_item(db, item_id)
        db.commit()
        return jsonify({"message": "重新計算完成", "unit_price": round(total, 2)})
    finally:
        db.close()


# ── 書籤（Bookmark）API ──

@app.route("/api/mrs-base/bookmarks", methods=["GET"])
@require_auth
def list_mrs_base_bookmarks(user_id):
    """取得我的書籤列表"""
    db = next(get_db())
    try:
        bookmarks = db.query(MrsBaseBookmark).filter(
            MrsBaseBookmark.user_id == user_id
        ).order_by(MrsBaseBookmark.created_at.desc()).all()
        result = []
        for bm in bookmarks:
            d = model_to_dict(bm)
            # 帶入項目基本資料
            item = db.query(MrsBaseItem).filter(MrsBaseItem.id == bm.item_id).first()
            if item:
                d["item"] = model_to_dict(item)
            result.append(d)
        return jsonify(result)
    finally:
        db.close()


@app.route("/api/mrs-base/bookmarks", methods=["POST"])
@require_auth
def create_mrs_base_bookmark(user_id):
    """新增書籤"""
    data = request.get_json()
    if not data or "item_id" not in data:
        return jsonify({"detail": "請提供 item_id"}), 400
    db = next(get_db())
    try:
        item_id = data["item_id"]
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404

        # 檢查是否已存在書籤
        existing = db.query(MrsBaseBookmark).filter(
            MrsBaseBookmark.user_id == user_id,
            MrsBaseBookmark.item_id == item_id
        ).first()
        if existing:
            return jsonify({"detail": "已加入書籤", "bookmark": model_to_dict(existing)}), 200

        bm = MrsBaseBookmark(
            user_id=user_id,
            item_id=item_id,
        )
        db.add(bm)
        db.commit()
        db.refresh(bm)
        return jsonify(model_to_dict(bm)), 201
    finally:
        db.close()


@app.route("/api/mrs-base/bookmarks/<int:bm_id>", methods=["DELETE"])
@require_auth
def delete_mrs_base_bookmark(bm_id, user_id):
    """移除書籤"""
    db = next(get_db())
    try:
        bm = db.query(MrsBaseBookmark).filter(
            MrsBaseBookmark.id == bm_id,
            MrsBaseBookmark.user_id == user_id
        ).first()
        if not bm:
            return jsonify({"detail": "書籤不存在"}), 404
        db.delete(bm)
        db.commit()
        return jsonify({"message": "書籤已移除"})
    finally:
        db.close()


# ── 搜尋 API ──

@app.route("/api/mrs-base/search", methods=["GET"])
@require_auth
def search_mrs_base(user_id):
    """模糊搜尋名稱/代碼（跨分類）"""
    db = next(get_db())
    try:
        q = request.args.get("q", "")
        category = request.args.get("category")
        kind = request.args.get("kind")

        query = db.query(MrsBaseItem)
        if q:
            pattern = f"%{q}%"
            query = query.filter(
                MrsBaseItem.code.ilike(pattern) | MrsBaseItem.c_name.ilike(pattern)
            )
        if category:
            # 支援 "," 分隔多個分類
            cat_ids = [int(x.strip()) for x in category.split(",") if x.strip()]
            if cat_ids:
                query = query.filter(MrsBaseItem.category_id.in_(cat_ids))
        if kind:
            query = query.filter(MrsBaseItem.cost_kind == kind)

        query = query.order_by(MrsBaseItem.code).limit(100)
        items = query.all()
        return jsonify([model_to_dict(i) for i in items])
    finally:
        db.close()


# ── 引用（Link to Budget）API ──

@app.route("/api/mrs-base/items/<int:item_id>/link-to-budget", methods=["POST"])
@require_auth
def link_mrs_base_to_budget(item_id, user_id):
    """將此單價引用到指定專案的預算項"""
    data = request.get_json()
    if not data or "project_id" not in data:
        return jsonify({"detail": "請提供 project_id"}), 400
    db = next(get_db())
    try:
        # 檢查 MrsBase 項目
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "公共單價項目不存在"}), 404

        project_id = data["project_id"]
        proj, err = _check_project_access(db, project_id, user_id)
        if err:
            return err

        # 若有指定 budget_item_id，更新該預算項
        budget_item_id = data.get("budget_item_id")
        if budget_item_id:
            bi = db.query(BudgetItem).filter(
                BudgetItem.id == budget_item_id,
                BudgetItem.project_id == project_id
            ).first()
            if not bi:
                return jsonify({"detail": "預算項目不存在"}), 404
            bi.mrs_base_item_id = item_id
            bi.c_name = item.c_name
            bi.c_unit = item.c_unit
            bi.unit_price = item.unit_price
            db.commit()
            db.refresh(bi)
            return jsonify({"message": "已連結到預算項目", "budget_item": model_to_dict(bi)})
        else:
            return jsonify({"detail": "請提供 budget_item_id"}), 400
    finally:
        db.close()


@app.route("/api/mrs-base/items/<int:item_id>/linked-projects", methods=["GET"])
@require_auth
def get_mrs_base_linked_projects(item_id, user_id):
    """列出引用此單價的專案/預算項"""
    db = next(get_db())
    try:
        item = db.query(MrsBaseItem).filter(MrsBaseItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "項目不存在"}), 404
        budget_items = db.query(BudgetItem).filter(
            BudgetItem.mrs_base_item_id == item_id
        ).all()
        result = []
        for bi in budget_items:
            proj = db.query(Project).filter(Project.id == bi.project_id).first()
            result.append({
                "project_id": bi.project_id,
                "project_name": proj.name if proj else "",
                "budget_item_id": bi.id,
                "budget_item_name": bi.c_name,
            })
        return jsonify(result)
    finally:
        db.close()


# ═══════════════════════════════════════════════
# Admin API（系統維護，僅 admin 角色可存取）
# ═══════════════════════════════════════════════

# ─── 使用者管理 ───

@app.route("/api/admin/users", methods=["GET"])
@require_admin
def admin_list_users(user_id):
    """使用者列表（分頁、搜尋、篩選）"""
    db = next(get_db())
    try:
        q = request.args.get("q", "")
        role = request.args.get("role", "")
        is_active = request.args.get("is_active", "")
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 20))

        query = db.query(User)
        if q:
            pattern = f"%{q}%"
            query = query.filter(
                User.username.ilike(pattern) |
                User.display_name.ilike(pattern) |
                User.email.ilike(pattern)
            )
        if role:
            query = query.filter(User.role == role)
        if is_active in ("true", "false"):
            query = query.filter(User.is_active == (is_active == "true"))

        total = query.count()
        users = query.order_by(User.created_at.desc()).offset(
            (page - 1) * per_page).limit(per_page).all()
        return jsonify({
            "users": [model_to_dict(u) for u in users],
            "total": total,
        })
    finally:
        db.close()


@app.route("/api/admin/users", methods=["POST"])
@require_admin
def admin_create_user(user_id):
    """管理員代建使用者"""
    data = request.get_json()
    if not data or not data.get("username") or not data.get("password"):
        return jsonify({"detail": "帳號與密碼為必填"}), 400
    db = next(get_db())
    try:
        # 檢查帳號是否已存在
        if db.query(User).filter(User.username == data["username"]).first():
            return jsonify({"detail": "帳號已存在"}), 409
        user = User(
            username=data["username"],
            password_hash=get_password_hash(data["password"]),
            display_name=data.get("display_name", data["username"]),
            email=data.get("email"),
            company=data.get("company"),
            department=data.get("department"),
            phone=data.get("phone"),
            role=data.get("role", UserRole.EDITOR.value),
            is_active=data.get("is_active", True),
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        return jsonify(model_to_dict(user)), 201
    finally:
        db.close()


@app.route("/api/admin/users/<int:uid>", methods=["GET"])
@require_admin
def admin_get_user(uid, user_id):
    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return jsonify({"detail": "使用者不存在"}), 404
        return jsonify(model_to_dict(user))
    finally:
        db.close()


@app.route("/api/admin/users/<int:uid>", methods=["PUT"])
@require_admin
def admin_update_user(uid, user_id):
    """更新使用者資料（含角色、啟用狀態）"""
    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return jsonify({"detail": "使用者不存在"}), 404
        data = request.get_json()
        if not data:
            return jsonify({"detail": "無更新資料"}), 400
        # 可更新欄位
        for field in ("display_name", "email", "company", "department", "phone", "role", "is_active"):
            if field in data:
                setattr(user, field, data[field])
        # 若提供密碼則更新
        if data.get("password"):
            user.password_hash = get_password_hash(data["password"])
        db.commit()
        db.refresh(user)
        return jsonify(model_to_dict(user))
    finally:
        db.close()


@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@require_admin
def admin_delete_user(uid, user_id):
    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return jsonify({"detail": "使用者不存在"}), 404
        if user.role == UserRole.ADMIN.value:
            return jsonify({"detail": "不可刪除管理員"}), 403
        db.delete(user)
        db.commit()
        return jsonify({"message": "使用者已刪除"})
    finally:
        db.close()


@app.route("/api/admin/users/<int:uid>/toggle-active", methods=["POST"])
@require_admin
def admin_toggle_user_active(uid, user_id):
    db = next(get_db())
    try:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return jsonify({"detail": "使用者不存在"}), 404
        user.is_active = not user.is_active
        db.commit()
        db.refresh(user)
        return jsonify(model_to_dict(user))
    finally:
        db.close()


@app.route("/api/admin/users/<int:uid>/change-role", methods=["POST"])
@require_admin
def admin_change_user_role(uid, user_id):
    db = next(get_db())
    try:
        data = request.get_json()
        if not data or "role" not in data:
            return jsonify({"detail": "請提供角色"}), 400
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return jsonify({"detail": "使用者不存在"}), 404
        user.role = data["role"]
        db.commit()
        db.refresh(user)
        return jsonify(model_to_dict(user))
    finally:
        db.close()


# ─── 系統參數 API ───

@app.route("/api/admin/params", methods=["GET"])
@require_admin
def admin_list_params(user_id):
    db = next(get_db())
    try:
        category = request.args.get("category", "")
        query = db.query(SystemParameter).order_by(SystemParameter.category, SystemParameter.sort_order)
        if category:
            query = query.filter(SystemParameter.category == category)
        params = query.all()
        return jsonify([model_to_dict(p) for p in params])
    finally:
        db.close()


@app.route("/api/admin/params", methods=["POST"])
@require_admin
def admin_create_param(user_id):
    data = request.get_json()
    if not data or not data.get("category") or not data.get("code"):
        return jsonify({"detail": "分類與代碼為必填"}), 400
    db = next(get_db())
    try:
        p = SystemParameter(
            category=data["category"],
            code=data["code"],
            c_name=data.get("c_name", ""),
            c_value=data.get("c_value"),
            c_default=data.get("c_default"),
            sort_order=data.get("sort_order", 0),
            is_active=data.get("is_active", True),
            memo=data.get("memo"),
        )
        db.add(p)
        db.commit()
        db.refresh(p)
        return jsonify(model_to_dict(p)), 201
    finally:
        db.close()


@app.route("/api/admin/params/<int:pid>", methods=["PUT"])
@require_admin
def admin_update_param(pid, user_id):
    db = next(get_db())
    try:
        p = db.query(SystemParameter).filter(SystemParameter.id == pid).first()
        if not p:
            return jsonify({"detail": "參數不存在"}), 404
        data = request.get_json()
        if not data:
            return jsonify({"detail": "無更新資料"}), 400
        for field in ("category", "code", "c_name", "c_value", "c_default", "sort_order", "is_active", "memo"):
            if field in data:
                setattr(p, field, data[field])
        db.commit()
        db.refresh(p)
        return jsonify(model_to_dict(p))
    finally:
        db.close()


@app.route("/api/admin/params/<int:pid>", methods=["DELETE"])
@require_admin
def admin_delete_param(pid, user_id):
    db = next(get_db())
    try:
        p = db.query(SystemParameter).filter(SystemParameter.id == pid).first()
        if not p:
            return jsonify({"detail": "參數不存在"}), 404
        db.delete(p)
        db.commit()
        return jsonify({"message": "參數已刪除"})
    finally:
        db.close()


# ─── 代碼表 API ───

@app.route("/api/admin/code-tables", methods=["GET"])
@require_admin
def admin_list_code_tables(user_id):
    db = next(get_db())
    try:
        tables = db.query(CodeTable).order_by(CodeTable.table_code).all()
        return jsonify([model_to_dict(t) for t in tables])
    finally:
        db.close()


@app.route("/api/admin/code-tables", methods=["POST"])
@require_admin
def admin_create_code_table(user_id):
    data = request.get_json()
    if not data or not data.get("table_code"):
        return jsonify({"detail": "代碼表識別碼為必填"}), 400
    db = next(get_db())
    try:
        if db.query(CodeTable).filter(CodeTable.table_code == data["table_code"]).first():
            return jsonify({"detail": "代碼表已存在"}), 409
        t = CodeTable(
            table_code=data["table_code"],
            table_name=data.get("table_name", ""),
            memo=data.get("memo"),
            is_active=data.get("is_active", True),
        )
        db.add(t)
        db.commit()
        db.refresh(t)
        return jsonify(model_to_dict(t)), 201
    finally:
        db.close()


@app.route("/api/admin/code-tables/<int:tid>", methods=["PUT"])
@require_admin
def admin_update_code_table(tid, user_id):
    db = next(get_db())
    try:
        t = db.query(CodeTable).filter(CodeTable.id == tid).first()
        if not t:
            return jsonify({"detail": "代碼表不存在"}), 404
        data = request.get_json()
        if not data:
            return jsonify({"detail": "無更新資料"}), 400
        for field in ("table_code", "table_name", "memo", "is_active"):
            if field in data:
                setattr(t, field, data[field])
        db.commit()
        db.refresh(t)
        return jsonify(model_to_dict(t))
    finally:
        db.close()


@app.route("/api/admin/code-tables/<int:tid>", methods=["DELETE"])
@require_admin
def admin_delete_code_table(tid, user_id):
    db = next(get_db())
    try:
        t = db.query(CodeTable).filter(CodeTable.id == tid).first()
        if not t:
            return jsonify({"detail": "代碼表不存在"}), 404
        db.delete(t)
        db.commit()
        return jsonify({"message": "代碼表已刪除"})
    finally:
        db.close()


# ─── 代碼項 API ───

def _build_code_item_tree(db, table_id, parent_id=None):
    """遞迴建立代碼項樹狀結構"""
    items = db.query(CodeItem).filter(
        CodeItem.table_id == table_id,
        CodeItem.parent_id == parent_id
    ).order_by(CodeItem.sort_order, CodeItem.code).all()
    result = []
    for item in items:
        d = model_to_dict(item)
        d["children"] = _build_code_item_tree(db, table_id, item.id)
        result.append(d)
    return result


@app.route("/api/admin/code-tables/<int:table_id>/items", methods=["GET"])
@require_admin
def admin_list_code_items(table_id, user_id):
    db = next(get_db())
    try:
        t = db.query(CodeTable).filter(CodeTable.id == table_id).first()
        if not t:
            return jsonify({"detail": "代碼表不存在"}), 404
        tree = _build_code_item_tree(db, table_id)
        return jsonify(tree)
    finally:
        db.close()


@app.route("/api/admin/code-tables/<int:table_id>/items", methods=["POST"])
@require_admin
def admin_create_code_item(table_id, user_id):
    data = request.get_json()
    if not data or not data.get("code") or not data.get("c_name"):
        return jsonify({"detail": "代碼與名稱為必填"}), 400
    db = next(get_db())
    try:
        t = db.query(CodeTable).filter(CodeTable.id == table_id).first()
        if not t:
            return jsonify({"detail": "代碼表不存在"}), 404
        item = CodeItem(
            table_id=table_id,
            parent_id=data.get("parent_id"),
            code=data["code"],
            c_name=data["c_name"],
            sort_order=data.get("sort_order", 0),
            is_active=data.get("is_active", True),
            ext_data=data.get("ext_data"),
            memo=data.get("memo"),
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item)), 201
    finally:
        db.close()


@app.route("/api/admin/code-items/<int:item_id>", methods=["PUT"])
@require_admin
def admin_update_code_item(item_id, user_id):
    db = next(get_db())
    try:
        item = db.query(CodeItem).filter(CodeItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "代碼項不存在"}), 404
        data = request.get_json()
        if not data:
            return jsonify({"detail": "無更新資料"}), 400
        for field in ("code", "c_name", "sort_order", "is_active", "parent_id", "ext_data", "memo"):
            if field in data:
                setattr(item, field, data[field])
        db.commit()
        db.refresh(item)
        return jsonify(model_to_dict(item))
    finally:
        db.close()


@app.route("/api/admin/code-items/<int:item_id>", methods=["DELETE"])
@require_admin
def admin_delete_code_item(item_id, user_id):
    db = next(get_db())
    try:
        item = db.query(CodeItem).filter(CodeItem.id == item_id).first()
        if not item:
            return jsonify({"detail": "代碼項不存在"}), 404
        # 遞迴刪除子項
        def _delete_children(parent_id):
            children = db.query(CodeItem).filter(CodeItem.parent_id == parent_id).all()
            for c in children:
                _delete_children(c.id)
                db.delete(c)
        _delete_children(item.id)
        db.delete(item)
        db.commit()
        return jsonify({"message": "代碼項已刪除"})
    finally:
        db.close()


# ─── 組織機構 API ───

def _build_org_tree(db, parent_id=None):
    """遞迴建立組織樹狀結構"""
    items = db.query(Organization).filter(
        Organization.parent_id == parent_id
    ).order_by(Organization.sort_order, Organization.code).all()
    result = []
    for org in items:
        d = model_to_dict(org)
        d["children"] = _build_org_tree(db, org.id)
        result.append(d)
    return result


@app.route("/api/admin/organizations", methods=["GET"])
@require_admin
def admin_list_orgs(user_id):
    db = next(get_db())
    try:
        tree = _build_org_tree(db)
        return jsonify(tree)
    finally:
        db.close()


@app.route("/api/admin/organizations", methods=["POST"])
@require_admin
def admin_create_org(user_id):
    data = request.get_json()
    if not data or not data.get("code") or not data.get("c_name"):
        return jsonify({"detail": "代碼與名稱為必填"}), 400
    db = next(get_db())
    try:
        if db.query(Organization).filter(Organization.code == data["code"]).first():
            return jsonify({"detail": "組織代碼已存在"}), 409
        org = Organization(
            parent_id=data.get("parent_id"),
            code=data["code"],
            c_name=data["c_name"],
            org_type=data.get("org_type", "部門"),
            sort_order=data.get("sort_order", 0),
            is_active=data.get("is_active", True),
            contact_person=data.get("contact_person"),
            contact_phone=data.get("contact_phone"),
            address=data.get("address"),
            memo=data.get("memo"),
        )
        db.add(org)
        db.commit()
        db.refresh(org)
        return jsonify(model_to_dict(org)), 201
    finally:
        db.close()


@app.route("/api/admin/organizations/<int:oid>", methods=["PUT"])
@require_admin
def admin_update_org(oid, user_id):
    db = next(get_db())
    try:
        org = db.query(Organization).filter(Organization.id == oid).first()
        if not org:
            return jsonify({"detail": "組織不存在"}), 404
        data = request.get_json()
        if not data:
            return jsonify({"detail": "無更新資料"}), 400
        for field in ("parent_id", "code", "c_name", "org_type", "sort_order",
                      "is_active", "contact_person", "contact_phone", "address", "memo"):
            if field in data:
                setattr(org, field, data[field])
        db.commit()
        db.refresh(org)
        return jsonify(model_to_dict(org))
    finally:
        db.close()


@app.route("/api/admin/organizations/<int:oid>", methods=["DELETE"])
@require_admin
def admin_delete_org(oid, user_id):
    db = next(get_db())
    try:
        org = db.query(Organization).filter(Organization.id == oid).first()
        if not org:
            return jsonify({"detail": "組織不存在"}), 404
        # 遞迴刪除子組織
        def _delete_children(parent_id):
            children = db.query(Organization).filter(Organization.parent_id == parent_id).all()
            for c in children:
                _delete_children(c.id)
                db.delete(c)
        _delete_children(org.id)
        db.delete(org)
        db.commit()
        return jsonify({"message": "組織已刪除"})
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 比較分析 API
# ═══════════════════════════════════════════════

def _flatten_budget_items(db: Session, project_id: int, scope: str = "leaf"):
    """將預算項目展平為列表（非樹狀）
    - scope="leaf": 只取葉節點 (kind not in B/Z)
    - scope="all": 取所有項目
    """
    items = db.query(BudgetItem).filter(
        BudgetItem.project_id == project_id
    ).all()
    if scope == "leaf":
        return [i for i in items if i.kind not in (BudgetItemKind.B, BudgetItemKind.Z)]
    return items


def _compare_budget_items_core(db: Session, project_a_id: int, project_b_id: int, scope: str = "leaf"):
    """核心比較邏輯：回傳比較結果 dict，供 API 與 Excel 匯出共用"""
    items_a = _flatten_budget_items(db, project_a_id, scope)
    items_b = _flatten_budget_items(db, project_b_id, scope)

    # 建立 dict key → item 對照（以 print_no 配對）
    def make_key(item):
        return (item.print_no or item.c_name or f"id:{item.id}")

    dict_a = {make_key(i): i for i in items_a}
    dict_b = {make_key(i): i for i in items_b}

    all_keys = set(dict_a.keys()) | set(dict_b.keys())

    proj_a = db.query(Project).filter(Project.id == project_a_id).first()
    proj_b = db.query(Project).filter(Project.id == project_b_id).first()

    results = []
    summary = {
        "added_count": 0,
        "removed_count": 0,
        "modified_count": 0,
        "unchanged_count": 0,
    }

    for key in all_keys:
        item_a = dict_a.get(key)
        item_b = dict_b.get(key)

        if item_a and not item_b:
            # 只在 A 存在
            a_data = {
                "quantity": float(item_a.quantity or 0),
                "unit_price": float(item_a.unit_price or 0),
                "amount": float(item_a.amount or 0),
            }
            results.append({
                "key": key,
                "c_name": item_a.c_name or "",
                "c_unit": item_a.c_unit or "",
                "a": a_data,
                "b": {"quantity": 0, "unit_price": 0, "amount": 0},
                "diff": {"quantity": 0, "unit_price": 0, "amount": 0},
                "diff_pct": {"quantity": None, "unit_price": None, "amount": None},
                "status": "removed",
            })
            summary["removed_count"] += 1
        elif not item_a and item_b:
            # 只在 B 存在
            b_data = {
                "quantity": float(item_b.quantity or 0),
                "unit_price": float(item_b.unit_price or 0),
                "amount": float(item_b.amount or 0),
            }
            results.append({
                "key": key,
                "c_name": item_b.c_name or "",
                "c_unit": item_b.c_unit or "",
                "a": {"quantity": 0, "unit_price": 0, "amount": 0},
                "b": b_data,
                "diff": {"quantity": 0, "unit_price": 0, "amount": 0},
                "diff_pct": {"quantity": None, "unit_price": None, "amount": None},
                "status": "added",
            })
            summary["added_count"] += 1
        else:
            # 兩者都有 → 計算差異
            def calc_diff(a_val, b_val):
                diff = b_val - a_val
                if a_val != 0:
                    pct = round((diff / a_val) * 100, 2)
                else:
                    pct = None  # N/A
                return diff, pct

            q_a = float(item_a.quantity or 0)
            q_b = float(item_b.quantity or 0)
            p_a = float(item_a.unit_price or 0)
            p_b = float(item_b.unit_price or 0)
            amt_a = float(item_a.amount or 0)
            amt_b = float(item_b.amount or 0)

            diff_q, pct_q = calc_diff(q_a, q_b)
            diff_p, pct_p = calc_diff(p_a, p_b)
            diff_a, pct_a = calc_diff(amt_a, amt_b)

            # 判定狀態：任一欄位差異 > 0.01 視為 modified
            is_modified = abs(diff_q) > 0.01 or abs(diff_p) > 0.01 or abs(diff_a) > 0.01

            results.append({
                "key": key,
                "c_name": item_a.c_name or item_b.c_name or "",
                "c_unit": item_a.c_unit or item_b.c_unit or "",
                "a": {"quantity": q_a, "unit_price": p_a, "amount": amt_a},
                "b": {"quantity": q_b, "unit_price": p_b, "amount": amt_b},
                "diff": {"quantity": round(diff_q, 2), "unit_price": round(diff_p, 2), "amount": round(diff_a, 2)},
                "diff_pct": {"quantity": pct_q, "unit_price": pct_p, "amount": pct_a},
                "status": "modified" if is_modified else "unchanged",
            })
            if is_modified:
                summary["modified_count"] += 1
            else:
                summary["unchanged_count"] += 1

    # 計算總計
    total_a = sum(r["a"]["amount"] for r in results)
    total_b = sum(r["b"]["amount"] for r in results)
    total_diff = total_b - total_a
    total_diff_pct = round((total_diff / total_a * 100), 2) if total_a != 0 else None

    return {
        "project_a": {"id": project_a_id, "name": proj_a.name if proj_a else ""},
        "project_b": {"id": project_b_id, "name": proj_b.name if proj_b else ""},
        "items": results,
        "summary": {
            "total_a": round(total_a, 2),
            "total_b": round(total_b, 2),
            "diff": round(total_diff, 2),
            "diff_pct": total_diff_pct,
            **summary,
        },
    }


@app.route("/api/compare/budget-items", methods=["POST"])
@require_auth
def compare_budget_items(user_id):
    """比較兩個專案的預算項目差異"""
    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供比較資料"}), 400

    project_a_id = data.get("project_a_id")
    project_b_id = data.get("project_b_id")
    scope = data.get("scope", "leaf")

    if not project_a_id or not project_b_id:
        return jsonify({"detail": "請提供 project_a_id 與 project_b_id"}), 400

    db = next(get_db())
    try:
        # 檢查專案存取權限
        _, err_a = _check_project_access(db, project_a_id, user_id)
        if err_a:
            return err_a
        _, err_b = _check_project_access(db, project_b_id, user_id)
        if err_b:
            return err_b

        result = _compare_budget_items_core(db, project_a_id, project_b_id, scope)
        return jsonify(result)
    finally:
        db.close()


@app.route("/api/compare/budget-items", methods=["GET"])
@require_auth
def compare_budget_items_get(user_id):
    """GET 版比較兩個專案的預算項目差異（透過查詢參數）"""
    project_a_id = request.args.get("project_a_id", type=int)
    project_b_id = request.args.get("project_b_id", type=int)
    scope = request.args.get("scope", "leaf")

    if not project_a_id or not project_b_id:
        return jsonify({"detail": "請提供 project_a_id 與 project_b_id"}), 400

    db = next(get_db())
    try:
        _, err_a = _check_project_access(db, project_a_id, user_id)
        if err_a:
            return err_a
        _, err_b = _check_project_access(db, project_b_id, user_id)
        if err_b:
            return err_b

        result = _compare_budget_items_core(db, project_a_id, project_b_id, scope=scope)
        return jsonify(result)
    finally:
        db.close()


@app.route("/api/compare/budget-items/export/excel", methods=["POST"])
@require_auth
def compare_budget_items_export_excel(user_id):
    """匯出工項比較報表 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    data = request.get_json()
    if not data:
        return jsonify({"detail": "請提供比較資料"}), 400

    project_a_id = data.get("project_a_id")
    project_b_id = data.get("project_b_id")

    if not project_a_id or not project_b_id:
        return jsonify({"detail": "請提供 project_a_id 與 project_b_id"}), 400

    db = next(get_db())
    try:
        _, err_a = _check_project_access(db, project_a_id, user_id)
        if err_a:
            return err_a
        _, err_b = _check_project_access(db, project_b_id, user_id)
        if err_b:
            return err_b

        result = _compare_budget_items_core(db, project_a_id, project_b_id)
        proj_a = result["project_a"]
        proj_b = result["project_b"]
        items = result["items"]
        summary = result["summary"]

        wb = Workbook()
        ws = wb.active
        ws.title = "工項比較報表"

        # 樣式
        title_font = Font(name="微軟正黑體", size=16, bold=True)
        sub_font = Font(name="微軟正黑體", size=11, bold=True)
        header_font = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
        normal_font = Font(name="微軟正黑體", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # 標題
        ws.merge_cells("A1:M1")
        ws["A1"] = f"比較報表 — {proj_a['name']} vs {proj_b['name']}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")

        # 統計摘要
        ws.merge_cells("A2:M2")
        diff_str = f"{summary['diff']:+,.2f}" if summary.get("diff") is not None else "N/A"
        ws["A2"] = (
            f"A 總額: {summary['total_a']:,.2f}  |  "
            f"B 總額: {summary['total_b']:,.2f}  |  "
            f"差異: {diff_str}  |  "
            f"新增: {summary['added_count']}  |  "
            f"移除: {summary['removed_count']}  |  "
            f"修改: {summary['modified_count']}  |  "
            f"不變: {summary['unchanged_count']}"
        )
        ws["A2"].font = sub_font
        ws["A2"].alignment = Alignment(horizontal="center")

        # 表頭
        headers = ["項次", "項目名稱", "單位",
                    "A 數量", "A 單價", "A 金額",
                    "B 數量", "B 單價", "B 金額",
                    "數量差異", "單價差異", "金額差異", "狀態"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 資料列
        for row_idx, item in enumerate(items, 5):
            data_row = [
                item["key"],
                item["c_name"],
                item["c_unit"],
                item["a"]["quantity"],
                item["a"]["unit_price"],
                item["a"]["amount"],
                item["b"]["quantity"],
                item["b"]["unit_price"],
                item["b"]["amount"],
            ]
            # 差異欄位
            for diff_field in ["quantity", "unit_price", "amount"]:
                d = item["diff"][diff_field]
                pct = item["diff_pct"][diff_field]
                if pct is not None:
                    data_row.append(f"{d:+,.2f} ({pct:+.2f}%)")
                else:
                    data_row.append(f"{d:+,.2f} (N/A)")
            data_row.append(item["status"])

            for col_idx, val in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right" if col_idx >= 4 else "left")

                # 標色
                status = item["status"]
                if status == "added":
                    cell.fill = green_fill
                elif status == "removed":
                    cell.fill = red_fill
                elif status == "modified":
                    # 差異欄位 (col 10~12) 若差異 > ±5% 標黃色
                    if col_idx in (10, 11, 12):
                        pct = item["diff_pct"][["quantity", "unit_price", "amount"][col_idx - 10]]
                        if pct is not None and abs(pct) > 5:
                            cell.fill = yellow_fill

        # 合計列
        total_row = len(items) + 5
        ws.cell(row=total_row, column=1, value="合計").font = sub_font
        ws.cell(row=total_row, column=6, value=summary["total_a"]).font = sub_font
        ws.cell(row=total_row, column=9, value=summary["total_b"]).font = sub_font
        ws.cell(row=total_row, column=12, value=f"{summary['diff']:+,.2f}").font = sub_font
        for col_idx in range(1, 14):
            ws.cell(row=total_row, column=col_idx).border = thin_border

        # 欄寬
        col_widths = [12, 25, 8, 12, 12, 14, 12, 12, 14, 18, 18, 18, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # 儲存至暫存檔（使用絕對路徑，避免 Flask send_file 路徑解析問題）
        safe_name_a = proj_a["name"][:10].replace("/", "_")
        safe_name_b = proj_b["name"][:10].replace("/", "_")
        filename = f"PCCES_比較報表_{safe_name_a}_vs_{safe_name_b}.xlsx"
        abs_report_dir = os.path.abspath(REPORT_DIR)
        filepath = os.path.join(abs_report_dir, filename)

        import tempfile as _tf
        os.makedirs(abs_report_dir, exist_ok=True)
        wb.save(filepath)

        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        db.close()


@app.route("/api/compare/mrs-base-prices", methods=["POST"])
@require_auth
def compare_mrs_base_prices(user_id):
    """MrsBase 單價比較 / 一覽 API"""
    data = request.get_json() or {}
    category_id = data.get("category_id")
    item_ids = data.get("item_ids")
    compare_type = data.get("compare_type", "all")

    db = next(get_db())
    try:
        query = db.query(MrsBaseItem)

        if category_id:
            query = query.filter(MrsBaseItem.category_id == category_id)
        if item_ids:
            query = query.filter(MrsBaseItem.id.in_(item_ids))

        query = query.order_by(MrsBaseItem.code)
        items = query.all()

        result_items = []
        prices = []
        for item in items:
            d = model_to_dict(item)
            # 帶入工料機組成
            breakdowns = db.query(MrsBaseBreakdownItem).filter(
                MrsBaseBreakdownItem.item_id == item.id
            ).order_by(MrsBaseBreakdownItem.id).all()
            d["breakdown_items"] = [model_to_dict(b) for b in breakdowns]
            d["has_analysis"] = item.is_analysis
            result_items.append(d)
            prices.append(item.unit_price or 0)

        avg_price = round(sum(prices) / len(prices), 2) if prices else 0
        max_price = max(prices) if prices else 0
        min_price = min(prices) if prices else 0

        return jsonify({
            "items": result_items,
            "summary": {
                "total": len(result_items),
                "avg_price": avg_price,
                "max_price": max_price,
                "min_price": min_price,
            },
        })
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 版本資訊 + 健康檢查 API（公開）
# ═══════════════════════════════════════════════


@app.route("/api/system/version")
def system_version():
    """取得系統版本資訊"""
    return jsonify({
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "build_date": BUILD_DATE,
        "repo_url": REPO_URL,
        "release_notes_url": RELEASE_NOTES_URL,
        "changelog": CHANGELOG,
        "dependencies": DEPENDENCIES,
    })


@app.route("/api/system/health")
def system_health():
    """系統健康檢查"""
    uptime_seconds = int(time.time() - START_TIME)
    db_status = "disconnected"
    try:
        db = next(get_db())
        db.execute(text("SELECT 1"))
        db.close()
        db_status = "connected"
    except Exception:
        db_status = "disconnected"

    return jsonify({
        "status": "healthy" if db_status == "connected" else "degraded",
        "database": db_status,
        "uptime_seconds": uptime_seconds,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })


# ═══════════════════════════════════════════════
# 功能開關 API（公開 — 已認證使用者可查詢啟用的開關）
# ═══════════════════════════════════════════════


@app.route("/api/feature-flags")
@require_auth
def list_enabled_feature_flags(user_id):
    """取得所有已啟用的功能開關（前端用於決定 UI 顯示）"""
    db = next(get_db())
    try:
        flags = db.query(FeatureFlag).filter(
            FeatureFlag.is_enabled == True
        ).order_by(FeatureFlag.sort_order).all()
        return jsonify([{
            "id": f.id,
            "flag_key": f.flag_key,
            "display_name": f.display_name,
            "description": f.description,
            "category": f.category,
            "is_enabled": f.is_enabled,
            "is_system": f.is_system,
            "sort_order": f.sort_order,
            "created_at": f.created_at.isoformat() if f.created_at else None,
            "updated_at": f.updated_at.isoformat() if f.updated_at else None,
        } for f in flags])
    finally:
        db.close()


# ═══════════════════════════════════════════════
# 功能開關管理 API（Admin Only）
# ═══════════════════════════════════════════════


@app.route("/api/admin/feature-flags", methods=["GET"])
@require_admin
def admin_list_feature_flags(user_id):
    """管理員：取得功能開關列表（支援分頁與分類篩選）"""
    db = next(get_db())
    try:
        query = db.query(FeatureFlag)

        category = request.args.get("category")
        if category and category != "all":
            query = query.filter(FeatureFlag.category == category)

        # 分頁
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)
        per_page = min(per_page, 200)

        total = query.count()
        flags = query.order_by(FeatureFlag.sort_order).offset(
            (page - 1) * per_page
        ).limit(per_page).all()

        return jsonify({
            "total": total,
            "flags": [{
                "id": f.id,
                "flag_key": f.flag_key,
                "display_name": f.display_name,
                "description": f.description,
                "category": f.category,
                "is_enabled": f.is_enabled,
                "is_system": f.is_system,
                "sort_order": f.sort_order,
                "created_at": f.created_at.isoformat() if f.created_at else None,
                "updated_at": f.updated_at.isoformat() if f.updated_at else None,
            } for f in flags],
        })
    finally:
        db.close()


@app.route("/api/admin/feature-flags", methods=["POST"])
@require_admin
def admin_create_feature_flag(user_id):
    """管理員：新增功能開關"""
    data = request.get_json() or {}
    flag_key = data.get("flag_key", "").strip()
    if not flag_key:
        return jsonify({"error": "功能代號不可為空"}), 400

    db = next(get_db())
    try:
        # 檢查是否已存在
        existing = db.query(FeatureFlag).filter(FeatureFlag.flag_key == flag_key).first()
        if existing:
            return jsonify({"error": f"功能代號 '{flag_key}' 已存在"}), 409

        flag = FeatureFlag(
            flag_key=flag_key,
            display_name=data.get("display_name", flag_key),
            description=data.get("description"),
            category=data.get("category", "general"),
            is_enabled=data.get("is_enabled", True),
            is_system=data.get("is_system", False),
            sort_order=data.get("sort_order", 0),
        )
        db.add(flag)
        db.commit()
        db.refresh(flag)
        return jsonify({
            "id": flag.id,
            "flag_key": flag.flag_key,
            "display_name": flag.display_name,
            "description": flag.description,
            "category": flag.category,
            "is_enabled": flag.is_enabled,
            "is_system": flag.is_system,
            "sort_order": flag.sort_order,
            "created_at": flag.created_at.isoformat() if flag.created_at else None,
            "updated_at": flag.updated_at.isoformat() if flag.updated_at else None,
        }), 201
    finally:
        db.close()


@app.route("/api/admin/feature-flags/<int:flag_id>", methods=["PUT"])
@require_admin
def admin_update_feature_flag(user_id, flag_id):
    """管理員：更新功能開關"""
    data = request.get_json() or {}
    db = next(get_db())
    try:
        flag = db.query(FeatureFlag).filter(FeatureFlag.id == flag_id).first()
        if not flag:
            return jsonify({"error": "功能開關不存在"}), 404

        # 可更新欄位（不可修改 flag_key）
        if "display_name" in data:
            flag.display_name = data["display_name"]
        if "description" in data:
            flag.description = data["description"]
        if "category" in data:
            flag.category = data["category"]
        if "sort_order" in data:
            flag.sort_order = data["sort_order"]
        if "is_enabled" in data:
            # 系統核心開關不可停用
            if flag.is_system and not data["is_enabled"]:
                return jsonify({"error": "系統核心功能不可停用"}), 403
            flag.is_enabled = data["is_enabled"]

        db.commit()
        db.refresh(flag)
        return jsonify({
            "id": flag.id,
            "flag_key": flag.flag_key,
            "display_name": flag.display_name,
            "description": flag.description,
            "category": flag.category,
            "is_enabled": flag.is_enabled,
            "is_system": flag.is_system,
            "sort_order": flag.sort_order,
            "created_at": flag.created_at.isoformat() if flag.created_at else None,
            "updated_at": flag.updated_at.isoformat() if flag.updated_at else None,
        })
    finally:
        db.close()


@app.route("/api/admin/feature-flags/<int:flag_id>", methods=["DELETE"])
@require_admin
def admin_delete_feature_flag(user_id, flag_id):
    """管理員：刪除功能開關（僅非 system 者可刪）"""
    db = next(get_db())
    try:
        flag = db.query(FeatureFlag).filter(FeatureFlag.id == flag_id).first()
        if not flag:
            return jsonify({"error": "功能開關不存在"}), 404
        if flag.is_system:
            return jsonify({"error": "系統核心功能不可刪除"}), 403

        db.delete(flag)
        db.commit()
        return jsonify({"message": "已刪除"})
    finally:
        db.close()


@app.route("/api/admin/feature-flags/<int:flag_id>/toggle", methods=["POST"])
@require_admin
def admin_toggle_feature_flag(user_id, flag_id):
    """管理員：切換功能開關啟用/停用"""
    db = next(get_db())
    try:
        flag = db.query(FeatureFlag).filter(FeatureFlag.id == flag_id).first()
        if not flag:
            return jsonify({"error": "功能開關不存在"}), 404
        if flag.is_system:
            return jsonify({"error": "系統核心功能不可停用"}), 403

        flag.is_enabled = not flag.is_enabled
        db.commit()
        db.refresh(flag)
        return jsonify({
            "id": flag.id,
            "flag_key": flag.flag_key,
            "display_name": flag.display_name,
            "description": flag.description,
            "category": flag.category,
            "is_enabled": flag.is_enabled,
            "is_system": flag.is_system,
            "sort_order": flag.sort_order,
            "created_at": flag.created_at.isoformat() if flag.created_at else None,
            "updated_at": flag.updated_at.isoformat() if flag.updated_at else None,
        })
    finally:
        db.close()


# ─── 前端靜態檔案服務 ───
# Vercel 會將 outputDirectory 的靜態檔案自動對應到根路徑
# 此 catch-all 僅處理 SPA fallback（其他靜態檔由 Vercel 直接 serve）
FRONTEND_DIST = os.path.join(os.path.dirname(__file__), 'static')


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_frontend(path):
    """提供前端靜態檔案，非 API 路徑一律回傳 index.html（SPA fallback）"""
    # API 路徑跳過
    if path and path.startswith("api/"):
        return jsonify({"detail": "Not found"}), 404
    # 先嘗試找實際檔案
    if path:
        file_path = os.path.join(FRONTEND_DIST, path)
        if os.path.isfile(file_path):
            return send_from_directory(FRONTEND_DIST, path)
    # SPA fallback：所有其他路徑回傳 index.html
    index_path = os.path.join(FRONTEND_DIST, "index.html")
    if os.path.exists(index_path):
        return send_from_directory(FRONTEND_DIST, "index.html")
    return jsonify({"detail": "App not built"}), 503


from api.seed_data import seed_demo_data

# 啟動時間（用於 uptime 計算）
import time
START_TIME = time.time()

# Vercel Serverless：在首次請求時初始化資料庫
@app.before_request
def _ensure_db():
    """確保資料庫表格存在 + 寫入示範資料（僅執行一次）"""
    if not hasattr(app, '_db_initialized'):
        init_db()
        db = next(get_db())
        try:
            seeded = seed_demo_data(db)
            if seeded:
                print("✅ 示範資料已寫入資料庫")
        finally:
            db.close()
        app._db_initialized = True


# ═══════════════════════════════════════════════
# 啟動
# ═══════════════════════════════════════════════

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)
