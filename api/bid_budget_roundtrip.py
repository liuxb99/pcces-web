"""Phase 4 electronic bid reverse import, preflight and round-trip lineage."""
from __future__ import annotations

import base64
import io
import json
from datetime import datetime, timezone
from uuid import uuid4
from xml.etree import ElementTree as ET

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook
from sqlalchemy import Column, DateTime, Integer, MetaData, String, Table, Text, select

metadata = MetaData()
bid_import_sessions = Table(
    "bid_import_sessions", metadata,
    Column("id", String(100), primary_key=True),
    Column("source_format", String(30), nullable=False),
    Column("format_version", String(20), nullable=False),
    Column("source_bid_project_code", String(100), nullable=False),
    Column("target_budget_project_code", String(100), nullable=False),
    Column("source_conversion_session_id", String(100), nullable=True),
    Column("status", String(20), nullable=False),
    Column("report_json", Text, nullable=False),
    Column("items_json", Text, nullable=False),
    Column("created_by", String(100), nullable=False),
    Column("created_at", DateTime(timezone=True), nullable=False),
    Column("row_version", Integer, nullable=False, default=1),
)

XLSX_HEADERS = ("來源工項ID", "工項編碼", "名稱", "單位", "數量", "單價", "金額")


def _xlsx_bytes(payload: str | bytes) -> bytes:
    if isinstance(payload, bytes):
        return payload
    text = payload.strip()
    if text.startswith("data:"):
        text = text.split(",", 1)[1]
    try:
        return base64.b64decode(text, validate=True)
    except Exception as exc:
        raise ValueError("XLSX payload must be base64 encoded") from exc


def parse_xlsx(payload: str | bytes) -> tuple[str, str, str, list[dict]]:
    binary = _xlsx_bytes(payload)
    try:
        workbook = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("invalid XLSX payload") from exc
    if "電子標單" not in workbook.sheetnames:
        raise ValueError("XLSX must contain 電子標單 worksheet")
    sheet = workbook["電子標單"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or tuple(str(value or "").strip() for value in rows[0][:7]) != XLSX_HEADERS:
        raise ValueError("XLSX electronic bid headers do not match canonical contract")

    project = ""
    source_version = ""
    for defined_name in workbook.defined_names.values():
        value = str(defined_name.attr_text or "").strip('"')
        if defined_name.name == "ProjectCode":
            project = value
        elif defined_name.name == "SourceVersion":
            source_version = value

    items: list[dict] = []
    for index, row in enumerate(rows[1:], start=1):
        values = list(row[:7]) + [None] * max(0, 7 - len(row))
        if not any(value not in (None, "") for value in values):
            continue
        lineage, code, name, unit, quantity, unit_price, amount = values[:7]
        lineage_text = str(lineage or f"ROW-{index}")
        items.append({
            "source_budget_item_id": lineage_text,
            "id": lineage_text,
            "code": str(code or "").strip().upper(),
            "name": str(name or "").strip(),
            "unit": str(unit or "").strip(),
            "quantity": str(quantity if quantity is not None else "0"),
            "unit_price": str(unit_price if unit_price is not None else "0"),
            "amount": str(amount if amount is not None else "0"),
        })
    # SourceVersion is retained in the workbook for audit; the existing tuple
    # contract returns format/version/project/items, so the exchange version is 1.0.
    _ = source_version
    return "XLSX", "1.0", project, items


def detect_and_parse(payload: str | bytes, hinted_format: str = "") -> tuple[str, str, str, list[dict]]:
    hinted = hinted_format.strip().upper()
    if hinted == "XLSX" or isinstance(payload, bytes):
        return parse_xlsx(payload)
    text = payload.strip()
    if hinted == "BID_JSON" or text.startswith("{"):
        data = json.loads(text)
        return "BID_JSON", "2.0", str(data.get("project_code", "")), list(data.get("items") or [])
    root = ET.fromstring(text)
    if root.tag == "PCCESBidExchange":
        rows = root.findall("./Items/Item")
        fmt, version = "XML_NEW", root.attrib.get("version", "2.0")
    elif root.tag == "PCCES":
        rows = root.findall("./Detail/Record")
        fmt, version = "XML_LEGACY", root.attrib.get("version", "1.0")
    else:
        raise ValueError("unsupported electronic bid format")
    project = root.findtext("./Header/ProjectCode", default="")
    items = []
    for index, row in enumerate(rows, start=1):
        items.append({
            "source_budget_item_id": row.findtext("SourceItemId", default=""),
            "id": row.findtext("SourceItemId", default=f"ROW-{index}"),
            "code": row.findtext("Code", default="").strip().upper(),
            "name": row.findtext("Name", default="").strip(),
            "unit": row.findtext("Unit", default="").strip(),
            "quantity": row.findtext("Quantity", default="0"),
            "unit_price": row.findtext("UnitPrice", default="0"),
            "amount": row.findtext("Amount", default="0"),
        })
    return fmt, version, project, items


def import_preflight(items: list[dict]) -> dict:
    errors, warnings, seen = [], [], set()
    if not items:
        errors.append({"code": "EMPTY_BID"})
    for index, item in enumerate(items):
        code = str(item.get("code", "")).strip().upper()
        if not code:
            errors.append({"code": "MISSING_ITEM_CODE", "index": index})
        elif code in seen:
            errors.append({"code": "DUPLICATE_ITEM_CODE", "item_code": code})
        seen.add(code)
        if not str(item.get("name", "")).strip():
            warnings.append({"code": "MISSING_ITEM_NAME", "index": index})
        if not str(item.get("source_budget_item_id", item.get("id", ""))).strip():
            warnings.append({"code": "MISSING_ROUNDTRIP_LINEAGE", "index": index})
    return {"errors": errors, "warnings": warnings, "error_count": len(errors), "warning_count": len(warnings), "can_continue": not errors}


class BidBudgetRoundTripService:
    def __init__(self, engine):
        self.engine = engine
        metadata.create_all(engine)

    def create(self, body: dict, actor: str) -> dict:
        target = str(body.get("target_budget_project_code", "")).strip()
        if not target:
            raise ValueError("target_budget_project_code is required")
        payload = body.get("payload", "")
        fmt, version, source_project, items = detect_and_parse(payload, str(body.get("format", "")))
        report = import_preflight(items)
        session_id = str(uuid4())
        now = datetime.now(timezone.utc)
        status = "READY" if report["can_continue"] else "BLOCKED"
        with self.engine.begin() as conn:
            conn.execute(bid_import_sessions.insert().values(
                id=session_id, source_format=fmt, format_version=version,
                source_bid_project_code=source_project, target_budget_project_code=target,
                source_conversion_session_id=str(body.get("source_conversion_session_id", "")).strip() or None,
                status=status, report_json=json.dumps(report, ensure_ascii=False, sort_keys=True),
                items_json=json.dumps(items, ensure_ascii=False, sort_keys=True),
                created_by=actor, created_at=now, row_version=1,
            ))
        return self.get(session_id)

    def get(self, session_id: str) -> dict:
        with self.engine.connect() as conn:
            row = conn.execute(select(bid_import_sessions).where(bid_import_sessions.c.id == session_id)).mappings().first()
        if not row:
            raise LookupError("bid import session not found")
        item = dict(row)
        item["report"] = json.loads(item.pop("report_json"))
        item["items"] = json.loads(item.pop("items_json"))
        item["created_at"] = item["created_at"].isoformat()
        item["round_trip_lineage"] = {
            "source_conversion_session_id": item["source_conversion_session_id"],
            "source_bid_project_code": item["source_bid_project_code"],
            "target_budget_project_code": item["target_budget_project_code"],
            "item_links": [{"source_budget_item_id": x.get("source_budget_item_id", x.get("id", "")), "imported_budget_item_id": x.get("id", "")} for x in item["items"]],
        }
        item["deep_link"] = f"/app/conversions/import?session={session_id}"
        return item


def build_bid_budget_roundtrip_blueprint(service: BidBudgetRoundTripService, resolve_user_id):
    bp = Blueprint("bid_budget_roundtrip", __name__, url_prefix="/api/conversions")

    @bp.post("/import-preflight")
    def preflight():
        if resolve_user_id() is None:
            return jsonify({"code": "UNAUTHORIZED"}), 401
        body = request.get_json(silent=True) or {}
        try:
            fmt, version, project, items = detect_and_parse(body.get("payload", ""), str(body.get("format", "")))
            return jsonify({"format": fmt, "format_version": version, "source_bid_project_code": project, "report": import_preflight(items), "items": items})
        except (ValueError, ET.ParseError, json.JSONDecodeError) as exc:
            return jsonify({"code": "INVALID_ARGUMENT", "detail": str(exc)}), 400

    @bp.post("/import-sessions")
    def create_session():
        actor = resolve_user_id()
        if actor is None:
            return jsonify({"code": "UNAUTHORIZED"}), 401
        try:
            return jsonify(service.create(request.get_json(silent=True) or {}, str(actor))), 201
        except (ValueError, ET.ParseError, json.JSONDecodeError) as exc:
            return jsonify({"code": "INVALID_ARGUMENT", "detail": str(exc)}), 400

    @bp.get("/import-sessions/<session_id>")
    def get_session(session_id: str):
        try:
            return jsonify(service.get(session_id))
        except LookupError as exc:
            return jsonify({"code": "NOT_FOUND", "detail": str(exc)}), 404

    return bp
